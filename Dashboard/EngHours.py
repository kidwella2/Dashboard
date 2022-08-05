
import ctypes
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from PyQt5.QtCore import QObject, pyqtSignal
import datetime
from pathlib import Path


class EngHrs(QObject):
    maximum = pyqtSignal(int)  # signals to communicate worker thread with main thread
    progress = pyqtSignal(int)
    finished = pyqtSignal()

    def EngHrs(self):
        today = datetime.date.today()                   # Initialize date as last sat (for Eng Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%y')

        data_folder = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/")
        fname = data_folder / "CN47N.XLSX"

        proj_dash = "C:/Users/Austin.Kidwell/Desktop/Project Dashboard R2.xlsx"
        #proj_dash = Path("//CPROME/Eng_Share/System Engineering/Visual Management/Test Folder/Project Dashboard R2.xlsx")

        eng_hours = "C:/Users/Austin.Kidwell/Desktop/EngHrs.xlsx"
        #eng_hours = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/EngHrs.xlsx")

        df = pd.read_excel(fname, sheet_name="Sheet1", header=[0])
        try:
            ActiveProjects = pd.read_excel(proj_dash, sheet_name="Projects", header=[0], usecols="A:G")
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {proj_dash}", "Info", 0)
            return

        df.columns = ("Plant", "Network", "Activity", "Status", "Project definition", "WBS element", "Activity desc.",
                    "Work center", "Confirmation", "Actual duration", "Earliest start date (basic)",
                    "Actual finish date", "Work", "Actual work", "Processing % of work", "Remaining work")


        COV_data = df.loc[(df['Plant'] == 'RA01')]          # group data with pandas dataframes
        DAV_data = df.loc[(df['Plant'] == 'RA02')]

        mask = ActiveProjects['Site'].str.contains('COV|DAV', na=False, regex=True)
        ActiveProjects = ActiveProjects[mask]
        #ActiveProjects = ActiveProjects[ActiveProjects['Ship Date'].isna()]     # Find null
        mask = ActiveProjects['Site'].str.contains('COV') #, na=False, regex=True)
        CovProjects = ActiveProjects[mask]
        mask = ActiveProjects['Site'].str.contains('DAV') #, na=False, regex=True)
        DavProjects = ActiveProjects[mask]

        mask = COV_data['WBS element'].str.contains('EN')
        COV_data = COV_data[mask]
        mask = DAV_data['WBS element'].str.contains('EN')
        DAV_data = DAV_data[mask]
        # print(COV_data)

        maskMech = COV_data['Work center'].str.contains('ENGMECH')
        COV_MEdata = COV_data[maskMech]
        maskMech = DAV_data['Work center'].str.contains('ENGMECH')
        DAV_MEdata = DAV_data[maskMech]

        # COV_MEdata.head()
        # print(COV_MEdata)

        maskElec = COV_data['Work center'].str.contains('ENGELEC')
        COV_EEdata = COV_data[maskElec]
        maskElec = DAV_data['Work center'].str.contains('ENGELEC')
        DAV_EEdata = DAV_data[maskElec]

        # COV_EEdata.head()
        # print(COV_EEdata)

        COV_MEdata.index = range(len(COV_MEdata))  # Reset index
        DAV_MEdata.index = range(len(DAV_MEdata))
        COV_EEdata.index = range(len(COV_EEdata))
        DAV_EEdata.index = range(len(DAV_EEdata))
        CovProjects.index = range(len(CovProjects))
        DavProjects.index = range(len(DavProjects))

        eng_data_cov = []
        eng_data_dav = []
        cov_proj_list = []
        dav_proj_list = []
        count = 0
        self.maximum.emit(2 * (len(CovProjects) + len(DavProjects)))
        for i in range(len(CovProjects)):           # get COV data for Eng Hours
            w1, w2, w3, w4 = '0', '0', '0', '0'
            for j in range(len(COV_MEdata)):
                if CovProjects['Project Number'][i] == COV_MEdata['Project definition'][j]:
                    w1 = COV_MEdata['Work'][j]
                    w2 = COV_MEdata['Actual work'][j]
                    break
            for k in range(len(COV_EEdata)):
                if CovProjects['Project Number'][i] == COV_EEdata['Project definition'][k]:
                    w3 = COV_EEdata['Work'][k]
                    w4 = COV_EEdata['Actual work'][k]
                    break
            eng_data_cov.append([CovProjects['Project Number'][i], CovProjects['Machine Code'][i],
                             CovProjects['Customer'][i], CovProjects['Description'][i], w1, w2, w3, w4,
                                 CovProjects['Ship Date'][i]])
            cov_proj_list.append(CovProjects['Project Number'][i])
            count += 1
            self.progress.emit(count)

        for i in range(len(DavProjects)):           # get DAV data for Eng Hours
            w1, w2, w3, w4 = '0', '0', '0', '0'
            for j in range(len(DAV_MEdata)):
                if DavProjects['Project Number'][i] == DAV_MEdata['Project definition'][j]:
                    w1 = DAV_MEdata['Work'][j]
                    w2 = DAV_MEdata['Actual work'][j]
                    break
            for k in range(len(DAV_EEdata)):
                if DavProjects['Project Number'][i] == DAV_EEdata['Project definition'][k]:
                    w3 = DAV_EEdata['Work'][k]
                    w4 = DAV_EEdata['Actual work'][k]
                    break
            eng_data_dav.append([DavProjects['Project Number'][i], DavProjects['Machine Code'][i],
                             DavProjects['Customer'][i], DavProjects['Description'][i], w1, w2, w3, w4,
                                 DavProjects['Ship Date'][i]])
            dav_proj_list.append(DavProjects['Project Number'][i])
            count += 1
            self.progress.emit(count)

        try:
            wb = openpyxl.load_workbook(eng_hours)  # Make excel sheet ready to edit
        except Exception:
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {eng_hours}", "Info", 0)
            return
        ws = wb['Budget Hours']
        ws["J1"] = last_sat
        total = []
        delete = []
        total1 = []
        delete1 = []

        if not pd.notna(cov_proj_list).all() or not pd.notna(dav_proj_list).all():
            #eng_data_cov[:,0] = eng_data_cov[eng_data_cov[:,0] == eng_data_cov[:,0]]
            #print(eng_data_cov[:,0])
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, "Project dashboard project is missing", "Info", 0)
            return

        i = 0
        while ws[f"A{i + 6}"].value is not None:            # Update Eng hours based on current
            if ws[f"B{i + 6}"].value == 'COV':              # Update COV info
                for j in range(len(eng_data_cov)):
                    if (ws[f"A{i + 6}"].value).strip() == (eng_data_cov[j][0]).strip():
                        ws[f'A{i + 6}'], ws[f'B{i + 6}'], ws[f'C{i + 6}'] = eng_data_cov[j][0], 'COV', eng_data_cov[j][1]
                        ws[f'D{i + 6}'], ws[f'E{i + 6}'] = eng_data_cov[j][2], eng_data_cov[j][3]
                        ws[f'F{i + 6}'], ws[f'G{i + 6}'] = eng_data_cov[j][4], eng_data_cov[j][5]
                        ws[f'I{i + 6}'], ws[f'J{i + 6}'] = eng_data_cov[j][6], eng_data_cov[j][7]
                        ws[f'O{i + 6}'] = eng_data_cov[j][8]
                    elif ws[f"A{i + 6}"].value > (eng_data_cov[j][0]).strip() and ws[f"A{i + 5}"].value == 'Order':
                        ws.insert_rows(idx=i + 6, amount=1)
                        ws[f'A{i + 6}'], ws[f'B{i + 6}'], ws[f'C{i + 6}'] = eng_data_cov[j][0], 'COV', eng_data_cov[j][1]
                        ws[f'D{i + 6}'], ws[f'E{i + 6}'] = eng_data_cov[j][2], eng_data_cov[j][3]
                        ws[f'F{i + 6}'], ws[f'G{i + 6}'] = eng_data_cov[j][4], eng_data_cov[j][5]
                        ws[f'I{i + 6}'], ws[f'J{i + 6}'] = eng_data_cov[j][6], eng_data_cov[j][7]
                        ws[f'O{i + 6}'] = eng_data_cov[j][8]
                    elif ws[f"A{i + 6}"].value < (eng_data_cov[j][0]).strip() < (ws[f"A{i + 7}"].value).strip():
                        ws.insert_rows(idx=i + 7, amount=1)
                        ws[f'A{i + 7}'], ws[f'B{i + 7}'], ws[f'C{i + 7}'] = eng_data_cov[j][0], 'COV', eng_data_cov[j][1]
                        ws[f'D{i + 7}'], ws[f'E{i + 7}'] = eng_data_cov[j][2], eng_data_cov[j][3]
                        ws[f'F{i + 7}'], ws[f'G{i + 7}'] = eng_data_cov[j][4], eng_data_cov[j][5]
                        ws[f'I{i + 7}'], ws[f'J{i + 7}'] = eng_data_cov[j][6], eng_data_cov[j][7]
                        ws[f'O{i + 7}'] = eng_data_cov[j][8]
                    elif ws[f"A{i + 6}"].value < (eng_data_cov[j][0]).strip() and ws[f"B{i + 7}"].value == 'DAV':
                        ws.insert_rows(idx=i + 7, amount=1)
                        ws[f'A{i + 7}'], ws[f'B{i + 7}'], ws[f'C{i + 7}'] = eng_data_cov[j][0], 'COV', eng_data_cov[j][1]
                        ws[f'D{i + 7}'], ws[f'E{i + 7}'] = eng_data_cov[j][2], eng_data_cov[j][3]
                        ws[f'F{i + 7}'], ws[f'G{i + 7}'] = eng_data_cov[j][4], eng_data_cov[j][5]
                        ws[f'I{i + 7}'], ws[f'J{i + 7}'] = eng_data_cov[j][6], eng_data_cov[j][7]
                        ws[f'O{i + 7}'] = eng_data_cov[j][8]
                total.append(eng_data_cov[j][0])
                if ws[f"A{i + 6}"].value not in cov_proj_list:
                    delete.append(ws[f"A{i + 6}"].value)
            elif ws[f"B{i + 6}"].value == 'DAV':    # Update DAV info
                for k in range(len(eng_data_dav)):
                    if (ws[f"A{i + 6}"].value).strip() == (eng_data_dav[k][0]).strip():
                        ws[f'A{i + 6}'], ws[f'B{i + 6}'], ws[f'C{i + 6}'] = eng_data_dav[k][0], 'DAV', eng_data_dav[k][1]
                        ws[f'D{i + 6}'], ws[f'E{i + 6}'] = eng_data_dav[k][2], eng_data_dav[k][3]
                        ws[f'F{i + 6}'], ws[f'G{i + 6}'] = eng_data_dav[k][4], eng_data_dav[k][5]
                        ws[f'I{i + 6}'], ws[f'J{i + 6}'] = eng_data_dav[k][6], eng_data_dav[k][7]
                        ws[f'O{i + 6}'] = eng_data_dav[k][8]
                    elif ws[f"A{i + 6}"].value > (eng_data_dav[k][0]).strip() and ws[f"B{i + 5}"].value == 'COV':
                        ws.insert_rows(idx=i + 6, amount=1)
                        ws[f'A{i + 6}'], ws[f'B{i + 6}'], ws[f'C{i + 6}'] = eng_data_dav[k][0], 'DAV', eng_data_dav[k][1]
                        ws[f'D{i + 6}'], ws[f'E{i + 6}'] = eng_data_dav[k][2], eng_data_dav[k][3]
                        ws[f'F{i + 6}'], ws[f'G{i + 6}'] = eng_data_dav[k][4], eng_data_dav[k][5]
                        ws[f'I{i + 6}'], ws[f'J{i + 6}'] = eng_data_dav[k][6], eng_data_dav[k][7]
                        ws[f'O{i + 6}'] = eng_data_dav[k][8]
                    elif ws[f"A{i + 6}"].value < (eng_data_dav[k][0]).strip() and ws[f"A{i + 7}"].value is None:
                        ws.insert_rows(idx=i + 7, amount=1)
                        ws[f'A{i + 7}'], ws[f'B{i + 7}'], ws[f'C{i + 7}'] = eng_data_dav[k][0], 'DAV', eng_data_dav[k][1]
                        ws[f'D{i + 7}'], ws[f'E{i + 7}'] = eng_data_dav[k][2], eng_data_dav[k][3]
                        ws[f'F{i + 7}'], ws[f'G{i + 7}'] = eng_data_dav[k][4], eng_data_dav[k][5]
                        ws[f'I{i + 7}'], ws[f'J{i + 7}'] = eng_data_dav[k][6], eng_data_dav[k][7]
                        ws[f'O{i + 7}'] = eng_data_dav[k][8]
                    elif ws[f"A{i + 6}"].value < (eng_data_dav[k][0]).strip() < (ws[f"A{i + 7}"].value).strip():
                        ws.insert_rows(idx=i + 7, amount=1)
                        ws[f'A{i + 7}'], ws[f'B{i + 7}'], ws[f'C{i + 7}'] = eng_data_dav[k][0], 'DAV', eng_data_dav[k][1]
                        ws[f'D{i + 7}'], ws[f'E{i + 7}'] = eng_data_dav[k][2], eng_data_dav[k][3]
                        ws[f'F{i + 7}'], ws[f'G{i + 7}'] = eng_data_dav[k][4], eng_data_dav[k][5]
                        ws[f'I{i + 7}'], ws[f'J{i + 7}'] = eng_data_dav[k][6], eng_data_dav[k][7]
                        ws[f'O{i + 7}'] = eng_data_dav[k][8]
                total1.append(eng_data_dav[k][0])
                if ws[f"A{i + 6}"].value not in dav_proj_list:
                    delete1.append(ws[f"A{i + 6}"].value)
            i += 1
            count += 1
            self.progress.emit(count)

        #print(delete)
        #print(delete1)
        for i in range(len(total) + 5, 5, -1):      # Remove project not in active projects sheet
            if ws[f"A{i}"].value in delete:
                ws.delete_rows(idx=i, amount=1)

        for i in range(len(total) + len(total1) + 5, len(total) + 5, -1):
            if ws[f"A{i}"].value in delete1:
                ws.delete_rows(idx=i, amount=1)

        # Handle formatting
        thin = Side(border_style="thin", color="000000")
        not_center = [1, 2, 3, 4, 5, 12, 13, 14]
        for row in range(6, len(eng_data_cov) + len(eng_data_dav) + 6):     # insert formula and format sheet
            ws[f'H{row}'], ws[f'K{row}'] = f'=F{row}-G{row}', f'=I{row}-J{row}'
            ws[f'L{row}'], ws[f'M{row}'] = f'=IFERROR(G{row}/F{row},"")', f'=IFERROR(J{row}/I{row},"")',
            ws[f'N{row}'] = f'=IFERROR((G{row}+J{row})/(F{row}+I{row}),"")'
            ws[f'L{row}'].number_format = '0%'
            ws[f'M{row}'].number_format = '0%'
            ws[f'N{row}'].number_format = '0%'
            for col in range(1, 15):  # A-N
                ws.cell(row=row, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(row=row, column=col).font = Font(size=11.5)
                if col not in not_center:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        try:
            wb.save(eng_hours)
        except Exception:
            wb.close()
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {eng_hours} to allow editing", "Info", 0)
            return
        wb.close()
        self.progress.emit(0)
        self.finished.emit()
        ctypes.windll.user32.MessageBoxW(0, f"{eng_hours} update complete", "Info", 0)

    def EngShp(self):
        data_folder = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/")
        fname = data_folder / "CN47N.XLSX"

        eng_hours = "C:/Users/Austin.Kidwell/Desktop/EngHrs.xlsx"
        #eng_hours = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/EngHrs.xlsx")
        as_ship = "C:/Users/Austin.Kidwell/Desktop/Eng Hours as Shipped.xlsx"
        #as_ship = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/Eng Hours as Shipped.xlsx")

        df = pd.read_excel(fname, sheet_name="Sheet1", header=[0])
        try:
            EngHrs = pd.read_excel(eng_hours, sheet_name="Budget Hours", header=[4])
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {eng_hours}", "Info", 0)
            return
        try:
            AsShip = pd.read_excel(as_ship, sheet_name="Budget Hours", header=[4])
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {as_ship}", "Info", 0)
            return

        df.columns = ("Plant", "Network", "Activity", "Status", "Project definition", "WBS element", "Activity desc.",
                      "Work center", "Confirmation", "Actual duration", "Earliest start date (basic)",
                      "Actual finish date", "Work", "Actual work", "Processing % of work", "Remaining work")

        COV_data = df.loc[(df['Plant'] == 'RA01')]
        DAV_data = df.loc[(df['Plant'] == 'RA02')]

        mask = COV_data['WBS element'].str.contains('EN')
        COV_data = COV_data[mask]
        mask = DAV_data['WBS element'].str.contains('EN')
        DAV_data = DAV_data[mask]

        maskTeco = COV_data['Status'].str.contains('TECO')
        COV_TCdata = COV_data[maskTeco]
        maskTeco = DAV_data['Status'].str.contains('TECO')
        DAV_TCdata = DAV_data[maskTeco]

        COV_TCdata.index = range(len(COV_TCdata))
        DAV_TCdata.index = range(len(DAV_TCdata))

        delete = []
        delete1 = []
        self.maximum.emit(6)
        self.progress.emit(1)
        for i in range(len(EngHrs)):
            for j in range(len(AsShip)):
                if EngHrs['Order'][i] == AsShip['Project Number'][j]:
                    if EngHrs['Order'][i] not in delete and EngHrs['COV/DAV'][i] == 'COV' == AsShip['Site'][j]:
                        delete.append(EngHrs['Order'][i])
                    elif EngHrs['Order'][i] not in delete1 and EngHrs['COV/DAV'][i] == 'DAV' == AsShip['Site'][j]:
                        delete1.append(EngHrs['Order'][i])
        #print(len(delete), delete)
        #print(len(delete1), delete1)

        self.progress.emit(2)
        wb = openpyxl.load_workbook(eng_hours)  # Make excel sheet ready to edit
        ws = wb['Budget Hours']

        #print(len(EngHrs))
        for i in range(len(EngHrs) + 5, 5, -1):
            if ws[f"A{i}"].value in delete or ws[f"A{i}"].value in delete1:
                ws.delete_rows(idx=i, amount=1)

        for row in range(6, len(EngHrs) - len(delete) - len(delete1) + 6):
            ws[f'H{row}'], ws[f'K{row}'] = f'=F{row}-G{row}', f'=I{row}-J{row}'
            ws[f'L{row}'], ws[f'M{row}'] = f'=IFERROR(G{row}/F{row},"")', f'=IFERROR(J{row}/I{row},"")'
            ws[f'N{row}'] = f'=IFERROR((G{row}+J{row})/(F{row}+I{row}),"")'

        try:
            wb.save(eng_hours)
        except Exception:
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {eng_hours} to allow editing", "Info", 0)
            return
        wb.close()

        self.progress.emit(3)
        EngHrs = pd.read_excel(eng_hours, sheet_name="Budget Hours", header=[4])

        teco_cov = []  # get teco info
        teco_dav = []
        for i in range(len(EngHrs)):
            for j in range(len(COV_TCdata)):
                if EngHrs['Order'][i] == COV_TCdata['Project definition'][j]:
                    teco_cov.append(EngHrs['Order'][i])
                    break
        self.progress.emit(4)
        for i in range(len(EngHrs)):
            for j in range(len(DAV_TCdata)):
                if EngHrs['Order'][i] == DAV_TCdata['Project definition'][j]:
                    teco_dav.append(EngHrs['Order'][i])
                    break
        #print(teco_cov, teco_dav)
        self.progress.emit(5)
        teco_no_d = []
        for i in range(len(teco_cov)):
            if 'D009' not in teco_cov[i]:
                teco_no_d.append((teco_cov[i], 'COV'))
        for i in range(len(teco_dav)):
            if 'D009' not in teco_dav[i]:
                teco_no_d.append((teco_dav[i], 'DAV'))
        #print(teco_no_d)

        wb = openpyxl.load_workbook(as_ship)  # Make excel sheet ready to edit
        ws = wb['Budget Hours']

        last_row = 2
        for row in ws:
            if not all([cell.value == None for cell in row]):
                last_row += 1
        #print(last_row)

        for i in range(len(teco_no_d)):
            for j in range(len(EngHrs)):
                if teco_no_d[i][0] == EngHrs['Order'][j]:
                    row = last_row + i + 1
                    ws.insert_rows(idx=row, amount=1)
                    try:
                        ship_date = datetime.datetime.strptime(str(EngHrs['Unnamed: 14'][j].date()), '%Y-%m-%d')
                    except Exception:
                        ship_date = None
                    ws[f'A{row}'], ws[f'B{row}'] = teco_no_d[i][1], EngHrs['Order'][j]
                    ws[f'C{row}'], ws[f'D{row}'] = EngHrs['Customer'][j], EngHrs['As Sold\nBudget'][j]
                    ws[f'E{row}'], ws[f'G{row}'] = EngHrs['To Date\nActuals'][j], EngHrs['As Sold Budget'][j]
                    ws[f'H{row}'], ws[f'J{row}'].value = EngHrs['To Date\nActuals.1'][j], ship_date
        # resize the table as data expands
        tab = ws.tables["Table2"]
        current_table_last = int(tab.ref[4:len(tab.ref)])
        new_table_last = last_row + len(teco_no_d)
        if new_table_last > current_table_last:
            tab.ref = f"A5:L{new_table_last}"
        # Handle formatting
        thin = Side(border_style="thin", color="000000")
        not_center = [1, 2, 3, 10, 11, 12]
        for row in range(last_row + 1, last_row + 1 + len(teco_no_d)):
            ws[f'F{row}'], ws[f'I{row}'] = f'=D{row}-E{row}', f'=G{row}-H{row}'
            ws[f'K{row}'] = f'=IFERROR(\'Budget Hours\'!E{row}/\'Budget Hours\'!D{row},"")'
            ws[f'L{row}'] = f'=IFERROR(\'Budget Hours\'!H{row}/\'Budget Hours\'!G{row},"")'
            ws[f'D{row}'].number_format, ws[f'E{row}'].number_format, ws[f'F{row}'].number_format = '0', '0', '0'
            ws[f'G{row}'].number_format, ws[f'H{row}'].number_format, ws[f'I{row}'].number_format = '0', '0', '0'
            ws[f'K{row}'].number_format, ws[f'L{row}'].number_format = '0%', '0%'
            ws.cell(row=row, column=10).alignment = Alignment(horizontal='right')
            ws[f'J{row}'].number_format = 'M/D/YYYY'
            for col in range(1, 13):  # A-L
                ws.cell(row=row, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.cell(row=row, column=col).font = Font(size=11.5)
                if col not in not_center:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        try:
            wb.save(as_ship)
        except Exception:
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {as_ship} to allow editing", "Info", 0)
            return
        wb.close()
        self.progress.emit(6)

        wb = openpyxl.load_workbook(eng_hours)  # Make excel sheet ready to edit
        ws = wb['Budget Hours']

        # Remove Eng Hours entry if sent to "as shipped"
        for i in range(len(EngHrs), 5, -1):
            if ws[f"A{i}"].value in teco_cov and ws[f"B{i}"].value == 'COV':
                ws.delete_rows(idx=i, amount=1)
            elif ws[f"A{i}"].value in teco_dav and ws[f"B{i}"].value == 'DAV':
                ws.delete_rows(idx=i, amount=1)

        # Reset row formulas after data shifts
        for row in range(6, len(EngHrs) - len(teco_cov) - len(teco_dav) + 6):
            ws[f'H{row}'], ws[f'K{row}'] = f'=F{row}-G{row}', f'=I{row}-J{row}'
            ws[f'L{row}'], ws[f'M{row}'] = f'=IFERROR(G{row}/F{row},"")', f'=IFERROR(J{row}/I{row},"")'
            ws[f'N{row}'] = f'=IFERROR((G{row}+J{row})/(F{row}+I{row}),"")'
        try:
            wb.save(eng_hours)
        except Exception:
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {eng_hours} to allow editing", "Info", 0)
            return
        wb.close()
        self.progress.emit(0)
        self.finished.emit()
        ctypes.windll.user32.MessageBoxW(0, f"{len(teco_no_d)} orders were added to \'as shipped\'", "Info", 0)
