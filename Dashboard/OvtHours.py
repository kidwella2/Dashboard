
import os
import ctypes
from shutil import copytree, rmtree
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from PyQt5.QtCore import QObject, pyqtSignal
import datetime
from pathlib import Path
import xlwings as xw
import pythoncom


class OvtHrs(QObject):
    maximum = pyqtSignal(int)  # signals to communicate worker thread with main thread
    progress = pyqtSignal(int)
    finished = pyqtSignal()

    def WTSSetupCov(self):
        today = datetime.date.today()                       # Initialize date variables
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        cur_year = sat.year
        sat2 = today - datetime.timedelta(idx - 6)
        next_sat = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m %d %y')
        next_sat_f = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%Y')
        next_sat_f2 = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%y')
        cur2year = sat2.year
        sat3 = today - datetime.timedelta(idx - 13)
        next3sat = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m %d %y')
        next3sat_f = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m/%d/%Y')
        next3sat_f2 = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m/%d/%y')
        cur3year = sat3.year

        wts_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/WE {last_sat}/")
        #wts_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/{cur_year} COV Timesheets/WE {last_sat}/")
        new_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/WE {next_sat}/")
        #new_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/{cur2year} COV Timesheets/WE {next_sat}/")
        if os.path.exists(new_folder):
            next_sat = next3sat
            next_sat_f = next3sat_f
            next_sat_f2 = next3sat_f2
            cur2year = cur3year
            new_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/WE {next_sat}/")
            #new_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/"
            #                  f"{cur2year} COV Timesheets/WE {next_sat}/")
        if os.path.exists(new_folder):
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Weekly Time Sheet WE {next_sat} already exists", "Info", 0)
            return
        else:
            try:
                copytree(wts_folder, new_folder)            #copy/paste folder
            except Exception:
                if os.path.exists(new_folder):
                    rmtree(new_folder)                      #remove folder
                    self.finished.emit()
                    ctypes.windll.user32.MessageBoxW(0, f"Close excel files in {wts_folder} to allow editing", "Info", 0)
                else:
                    self.finished.emit()
                    ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate the folder {wts_folder}", "Info", 0)
                return

        target_files = []
        for root, dirs, files in os.walk(new_folder):  # Obtain files to reset data
            for file in files:
                if last_sat in file:
                    file_new = file.replace(f"{last_sat}", f"{next_sat}")
                    if "SUMMARY" not in file_new and "summary" not in file_new:
                        ftarget = os.path.join(root, file_new)
                        target_files.append(ftarget)
                        os.rename(os.path.join(root, file), ftarget)
                    else:
                        starget = os.path.join(root, file_new)
                        os.rename(os.path.join(root, file), starget)

        count = 0
        self.maximum.emit(len(target_files))
        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        excel_app.api.EnableEvents = False
        for file in target_files:                                   # Setup all but summary file
            wb = xw.Book(file)  # Make excel sheet ready to edit
            #print(file)
            for sheet in wb.sheets:
                sheet = sheet.name
                ws = wb.sheets[sheet]
                #print(sheet)
                #if '^' in sheet:
                    #ws.range("K47").clear_contents()
                if sheet != "Setup":               # Reset sheet info
                    ws.range("C6:I11").clear_contents()
                    ws.range("C14:I17").clear_contents()
                    ws.range("C20:I40").clear_contents()
                    ws.range("K6:K9").value = ''
                    ws.range("K10:K11").value = "Reason:"
                    ws.range("K14:M17").clear_contents()
                    for i in range(4):
                        if ws.range(f"A{i + 14}").value is None:
                            print('skip')
                        elif ws.range(f"A{i + 14}").value != '':
                            ws.range(f"A{i + 14}").value = ws.range(f"A{i + 14}").value[:5]
                else:
                    ws.range('B1').value = next_sat_f
                count += 1
                self.progress.emit(count)
            wb.save()
            wb.close()
        excel_app.kill()

        summ_file = new_folder / f"ENGINEERING TIMESHEET SUMMARY WITH BREAKOUT WE {next_sat}.xlsm"
        # print(summ_file)
        excel_app = xw.App(visible=False)
        wb = xw.Book(summ_file)
        wb.sheets['Summary']["B1"].value = f"{next_sat_f}"                  # set summary tab for new week
        wb.sheets['Summary']["B3"].value = f"{next_sat_f}"
        wb.sheets['Summary'].charts[0].name = 'Chart 1'
        wb.sheets['Summary'].charts['Chart 1'].api[1].ChartTitle.Text = f"Engineering Overtime - {next_sat_f2}"
        ws = wb.sheets["SAP Hours"]                                     # clear sap hours info
        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        #print(max_row)
        ws.range(f"A2:G{max_row}").clear_contents()
        wb.save()
        wb.close()
        excel_app.kill()
        junk_file = new_folder / "Thumbs.db"
        if os.path.exists(junk_file):
            os.remove(junk_file)
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"Setup for Weekly Time Sheet WE {next_sat} complete", "Info", 0)

    def WTSSetupDav(self):
        today = datetime.date.today()                       # Initialize date variables
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        last_sat_s = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m_%d_%y')
        cur_year = sat.year
        sat2 = today - datetime.timedelta(idx - 6)
        next_sat = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m %d %y')
        next_sat_s = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m_%d_%y')
        next_sat_f = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%Y')
        next_sat_f2 = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%y')
        cur2year = sat2.year
        sat3 = today - datetime.timedelta(idx - 13)
        next3sat = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m %d %y')
        next3sat_s = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m_%d_%y')
        next3sat_f = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m/%d/%Y')
        next3sat_f2 = datetime.datetime.strptime(str(sat3), '%Y-%m-%d').strftime('%m/%d/%y')
        cur3year = sat3.year

        wts_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {last_sat}/")
        #wts_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/{cur_year} DAV Timesheets/WE {last_sat}/")
        new_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {next_sat}/")
        #new_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/{cur2year} DAV Timesheets/WE {next_sat}/")
        if os.path.exists(new_folder):
            next_sat = next3sat
            next_sat_s = next3sat_s
            next_sat_f = next3sat_f
            next_sat_f2 = next3sat_f2
            cur2year = cur3year
            new_folder = Path(f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {next_sat}/")
            #new_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/"
            #                  f"{cur2year} DAV Timesheets/WE {next_sat}/")
        if os.path.exists(new_folder):
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Weekly Time Sheet WE {next_sat} already exists", "Info", 0)
            return
        else:
            try:
                copytree(wts_folder, new_folder)                            # copy/paste folder
            except Exception:
                if os.path.exists(new_folder):
                    rmtree(new_folder)                                      # remove folder
                    self.finished.emit()
                    ctypes.windll.user32.MessageBoxW(0, f"Close excel files in {wts_folder} to allow editing", "Info", 0)
                else:
                    self.finished.emit()
                    ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate the folder {wts_folder}", "Info", 0)
                return

        target_files = []
        for root, dirs, files in os.walk(new_folder):  # Obtain files to reset data
            for file in files:
                if last_sat in file:
                    file_new = file.replace(f"{last_sat}", f"{next_sat}")
                    if "SUMMARY" not in file_new and "summary" not in file_new:
                        ftarget = os.path.join(root, file_new)
                        target_files.append(ftarget)
                        os.rename(os.path.join(root, file), ftarget)
                elif last_sat_s in file:
                    file_summ = file.replace(f"{last_sat_s}", f"{next_sat_s}")
                    starget = os.path.join(root, file_summ)
                    os.rename(os.path.join(root, file), starget)

        count = 0
        self.maximum.emit(len(target_files))
        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        excel_app.api.EnableEvents = False
        for file in target_files:                                   # Setup all but summary file
            wb = xw.Book(file)  # Make excel sheet ready to edit
            #print(file)
            for sheet in wb.sheets:
                sheet = sheet.name
                ws = wb.sheets[sheet]
                #print(sheet)
                if '^' in sheet:
                    ws.range("K47").clear_contents()
                if sheet != "Setup":               # Reset sheet info
                    ws.range("C6:I11").clear_contents()                 # clear timesheet data
                    ws.range("C14:I17").clear_contents()
                    ws.range(f"C20:I40").clear_contents()
                    ws.range("K6:K9").value = ''
                    ws.range("K10:K11").value = "Reason:"
                    ws.range("K14:M17").clear_contents()
                    for i in range(4):
                        if ws.range(f"A{i + 14}").value is None:
                            print('skip')
                        elif len(ws.range(f"A{i + 14}").value) > 5:
                            ws.range(f"A{i + 14}").value = ws.range(f"A{i + 14}").value[:5]
                else:
                    ws.range('B1').value = next_sat_f
                count += 1
                self.progress.emit(count)
            wb.save()
            wb.close()
        excel_app.kill()

        summ_file = new_folder / f"ENGINEERING TIMESHEET SUMMARY WE {next_sat_s}_WithBreakout.xlsm"
        #print(summ_file)
        excel_app = xw.App(visible=False)
        wb = xw.Book(summ_file)
        wb.sheets['Summary']["B1"].value = f"{next_sat_f}"                      # setup summary tab
        wb.sheets['Summary']["B3"].value = f"{next_sat_f}"
        wb.sheets['Summary'].charts[0].name = 'Chart 1'
        wb.sheets['Summary'].charts['Chart 1'].api[1].ChartTitle.Text = f"Engineering Overtime - {next_sat_f2}"
        ws = wb.sheets["SAP Hours"]                                             # clear sap hours tab
        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        #print(max_row)
        ws.range(f"A2:G{max_row}").clear_contents()
        wb.save()
        wb.close()
        excel_app.kill()
        junk_file = new_folder / "Thumbs.db"
        if os.path.exists(junk_file):
            os.remove(junk_file)
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"Setup for Weekly Time Sheet WE {next_sat} complete", "Info", 0)

    def SetupCov(self):
        today = datetime.date.today()           # Initialize date variables
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        cur_year = sat.year
        sat2 = today - datetime.timedelta(idx - 6)
        next_sat = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m %d %y')
        next_sat_f = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%y')
        cur2year = sat2.year

        ot_hours = f"C:/Users/Austin.Kidwell/Desktop/Overtime COV {cur2year}.xlsx"
        #ot_hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
        #                f"overtime cov {cur2year}.xlsx")

        self.maximum.emit(1)
        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        if cur_year != cur2year:            # setup new wb if year changes
            ot2hours = f"C:/Users/Austin.Kidwell/Desktop/Overtime COV {cur_year}.xlsx"
            #ot2hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
            #                f"overtime cov {cur_year}.xlsx")
            try:
                wb1 = xw.Book(ot2hours)
            except Exception as e:  # must create sheet before populating it
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot2hours}", "Info", 0)
                return 1
            ws1 = wb1.sheets(1)
            if os.path.exists(ot_hours):
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Excel sheet WE {next_sat} already exists", "Info", 0)
                return
            wb2 = xw.Book()
            ws1.api.Copy(Before=wb2.sheets(1).api)
            wb2.sheets(2).delete()
            wb2.sheets(1).name = f'WE {next_sat}'
            wb2.sheets(1)["A1"].value = f"Engineering Overtime - WE {next_sat_f}"
            wb2.save(ot_hours)
            self.progress.emit(1)
        else:                               # setup new ws for the week
            try:
                wb1 = xw.Book(ot_hours)
            except Exception as e:  # must create sheet before populating it
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot_hours}", "Info", 0)
                return 1
            ws1 = wb1.sheets(1)
            if ws1.name == f'WE {next_sat}':
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Excel sheet WE {next_sat} already exists", "Info", 0)
                return
            ws1.api.Copy(Before=wb1.sheets(1).api)
            wb1.sheets(1).name = f'WE {next_sat}'
            wb1.sheets(1)["A1"].value = f"Engineering Overtime - WE {next_sat_f}"
            self.progress.emit(1)
            try:
                wb1.save(ot_hours)
            except Exception:
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Close excel file {ot_hours} to allow editing", "Info", 0)
                return
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"Setup for Overtime COV {cur2year} complete", "Info", 0)

    def SetupDav(self):
        today = datetime.date.today()  # Initialize date variables
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m-%d-%y')
        cur_year = sat.year
        sat2 = today - datetime.timedelta(idx - 6)
        next_sat = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m-%d-%y')
        next_sat_f = datetime.datetime.strptime(str(sat2), '%Y-%m-%d').strftime('%m/%d/%y')
        cur2year = sat2.year

        ot_hours = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/Overtime {cur2year}.xlsx"
        #ot_hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
        #                f"overtime {cur2year}.xlsx")

        self.maximum.emit(1)
        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        if cur_year != cur2year:  # setup new wb if year changes
            ot2hours = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/Overtime {cur_year}.xlsx"
            #ot2hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
            #                f"overtime {cur_year}.xlsx")
            try:
                wb1 = xw.Book(ot2hours)
            except Exception as e:  # must create sheet before populating it
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot2hours}", "Info", 0)
                return 1
            ws1 = wb1.sheets(len(wb1.sheets))
            if os.path.exists(ot_hours):
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Excel sheet {next_sat} already exists", "Info", 0)
                return
            wb2 = xw.Book()
            ws1.api.Copy(Before=wb2.sheets(1).api)
            wb2.sheets(2).delete()
            wb2.sheets(1).name = f'{next_sat}'
            wb2.sheets(1)["A1"].value = f"Engineering Overtime - WE {next_sat_f}"
            wb2.save(ot_hours)
            self.progress.emit(1)
        else:  # setup new ws for the week
            try:
                wb1 = xw.Book(ot_hours)
            except Exception as e:  # must create sheet before populating it
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot_hours}", "Info", 0)
                return 1
            ws1 = wb1.sheets(len(wb1.sheets))
            if ws1.name == f'{next_sat}':
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Excel sheet {next_sat} already exists", "Info", 0)
                return
            ws1.api.Copy(After=wb1.sheets(len(wb1.sheets)).api)
            wb1.sheets(len(wb1.sheets)).name = f'{next_sat}'
            wb1.sheets(len(wb1.sheets))["A1"].value = f"Engineering Overtime - WE {next_sat_f}"
            self.progress.emit(1)
            try:
                wb1.save(ot_hours)
            except Exception:
                excel_app.kill()
                self.progress.emit(0)
                self.finished.emit()
                pythoncom.CoUninitialize()
                ctypes.windll.user32.MessageBoxW(0, f"Close excel file {ot_hours} to allow editing", "Info", 0)
                return
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"Setup for Overtime {cur2year} complete", "Info", 0)

    def OvtHrsCov(self):
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        last_sat_f = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%Y')
        last_sat_f2 = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/WE {last_sat}/"
        #data_folder = f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/" \
        #              f"{cur_year} COV Timesheets/WE {last_sat}/"

        target_files = []
        for root, dirs, files in os.walk(data_folder):      # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))
        if target_files == []:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {data_folder}", "Info", 0)
            return
        ot_hours = f"C:/Users/Austin.Kidwell/Desktop/Overtime COV {cur_year}.xlsx"
        #ot_hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
        #                f"overtime cov {cur_year}.xlsx")

        ovt = []
        EE_total, EE_ovt, EE_ovt_paid, ME_total, ME_ovt, ME_ovt_paid = 0, 0, 0, 0, 0, 0
        EE_C, EE_R, EE_U, EE_D, ME_C, ME_R, ME_U, ME_D = 0, 0, 0, 0, 0, 0, 0, 0
        EE_admin, EE_vacation, EE_holiday, EE_sick, EE_funeral, EE_other, EE_service = 0, 0, 0, 0, 0, 0, 0
        ME_admin, ME_vacation, ME_holiday, ME_sick, ME_funeral, ME_other, ME_service = 0, 0, 0, 0, 0, 0, 0

        count = 0
        self.maximum.emit((2 * len(target_files))+2)
        for file in target_files:
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if "EE" in file and sheet != "Setup":               # Obtain EE info
                    ws = wb[sheet]
                    EE_total += ws["J41"].value
                    EE_vacation += ws["J6"].value
                    EE_holiday += ws["J7"].value
                    EE_sick += ws["J8"].value
                    EE_funeral += ws["J9"].value
                    EE_other += ws["J10"].value
                    EE_admin += ws["J11"].value
                    if ws["J42"].value != 0:
                        if "EE_COOP" in file or "ee_coop" in file:
                            type = "EE Coop"
                        elif "EE_CONTRACTOR" in file or "ee_contractor" in file:
                            type = "EE Contractor"
                        else:
                            type = "Electrical"
                        EE_ovt += ws["J42"].value
                        EE_ovt_paid += ws["J43"].value
                        ovt_list = [ws["C3"].value, ws["J41"].value, ws["J42"].value, ws["J43"].value, type]
                        ovt.append(ovt_list)
                    for s in range(4):
                        if ws[f"J{14 + s}"].value != 0:
                            EE_service += ws[f"J{14 + s}"].value
                    for i in range(21):
                        if ws[f"J{20 + i}"].value != 0:
                            char = (ws[f"A{20 + i}"].value[0]).upper()
                            if char == "C":
                                EE_C += ws[f"J{20 + i}"].value
                            elif char == "R":
                                EE_R += ws[f"J{20 + i}"].value
                            elif char == "U":
                                EE_U += ws[f"J{20 + i}"].value
                            elif char == "D":
                                EE_D += ws[f"J{20 + i}"].value
                elif "ME" in file and sheet != "Setup":               # Obtain ME info
                    ws = wb[sheet]
                    ME_total += ws["J41"].value
                    ME_vacation += ws["J6"].value
                    ME_holiday += ws["J7"].value
                    ME_sick += ws["J8"].value
                    ME_funeral += ws["J9"].value
                    ME_other += ws["J10"].value
                    ME_admin += ws["J11"].value
                    if ws["J42"].value != 0:
                        if "ME_COOP" in file or "me_coop" in file:
                            type = "ME Coop"
                        elif "ME_CONTRACTOR" in file or "me_contractor" in file:
                            type = "ME Contractor"
                        else:
                            type = "Mechanical"
                        ME_ovt += ws["J42"].value
                        ME_ovt_paid += ws["J43"].value
                        ovt_list = [ws["C3"].value, ws["J41"].value, ws["J42"].value, ws["J43"].value, type]
                        ovt.append(ovt_list)
                    for s in range(4):
                        if ws[f"J{14 + s}"].value != 0:
                            ME_service += ws[f"J{14 + s}"].value
                    for i in range(21):
                        if ws[f"J{20 + i}"].value != 0:
                            char = (ws[f"A{20 + i}"].value[0]).upper()
                            if char == "C":
                                ME_C += ws[f"J{20 + i}"].value
                            elif char == "R":
                                ME_R += ws[f"J{20 + i}"].value
                            elif char == "U":
                                ME_U += ws[f"J{20 + i}"].value
                            elif char == "D":
                                ME_D += ws[f"J{20 + i}"].value
            count += 1
            self.progress.emit(count)

        EE_sum = EE_C+EE_R+EE_U+EE_D+EE_admin+EE_vacation+EE_holiday+EE_sick+EE_funeral+EE_other+EE_service
        ME_sum = ME_C+ME_R+ME_U+ME_D+ME_admin+ME_vacation+ME_holiday+ME_sick+ME_funeral+ME_other+ME_service

        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        try:
            wb = xw.Book(ot_hours)                                      # Make excel sheet ready to edit (Yearly Overtime)
        except Exception:              # make sure yearly overtime file exist
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot_hours}", "Info", 0)
            return
        try:
            ws1 = wb.sheets[f'WE {last_sat}']
        except Exception as e:              # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Setup excel sheet WE {last_sat} first", "Info", 0)
            print('Catch: ', e.__class__)
            return 1
        last_no_ovt = 43
        num_ovt = len(ovt)
        this_last = last_no_ovt + num_ovt
        max_row = ws1.range('A' + str(ws1.cells.last_cell.row)).end('up').row
        move = this_last - max_row

        # Shift the data in the file depending on number of overtimes
        if move != 0:
            wb.sheets[f'WE {last_sat}'].range(f"A{num_ovt + 3 - move}:F{max_row}").copy(wb.sheets[f'WE {last_sat}']
                                                                             .range(f"A{num_ovt + 3}:F{this_last}"))
            ws1.charts["Chart 1"].delete()          # Fix issues from moving a chart
            ws1.charts[0].name = "Chart 1"
            ws1.charts["Chart 1"].set_source_data(ws1.range(f"B{num_ovt + 4}:E{num_ovt + 5}"))
        if move < 0:                            # Clear unneccessary data from moving content
            ws1.range(f"A{this_last + 1}:F{max_row}").clear() # for shrink
        elif move > 0:
            ws1.range(f"B{num_ovt - move + 3}:E{num_ovt - move + 4}").clear()  # for grow

        if num_ovt > 0:
            for border_id in range(7, 13):                          # Reformat the table containing overtime employees
                ws1.range(f"A3:E{num_ovt + 2}").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A3:E{num_ovt + 2}").api.Borders(border_id).Weight = 2
        if move > 0:
            ws1.range(f"A3:D{num_ovt + 2}").api.Font.Size = 10
            ws1.range(f"A3:D{num_ovt + 2}").api.Font.Name = "Arial"

        for i in range(num_ovt):                                    # Populate the overtime table at the top
            ws1.cells(i + 3, 1).value = ovt[i][0]
            ws1.cells(i + 3, 2).value = int(ovt[i][1] + 0.5)
            ws1.cells(i + 3, 3).value = int(ovt[i][2] + 0.5)
            ws1.cells(i + 3, 4).value = int(ovt[i][3] + 0.5)
            ws1.cells(i + 3, 5).value = ovt[i][4]

        # Populate new data for current week
        ws1.range(f"B{num_ovt + 5}").value = int(ME_total + 0.5)
        ws1.range(f"C{num_ovt + 5}").value = int(ME_ovt + 0.5)
        ws1.range(f"D{num_ovt + 5}").value = int(EE_total + 0.5)
        ws1.range(f"E{num_ovt + 5}").value = int(EE_ovt + 0.5)

        ws1.range(f"B{this_last - 11}").value = int(ME_C + 0.5)
        ws1.range(f"B{this_last - 10}").value = int(ME_R + 0.5)
        ws1.range(f"B{this_last - 9}").value = int(ME_U + 0.5)
        ws1.range(f"B{this_last - 8}").value = int(ME_D + 0.5)
        ws1.range(f"B{this_last - 7}").value = int(ME_admin + 0.5)
        ws1.range(f"B{this_last - 6}").value = int(ME_vacation + 0.5)
        ws1.range(f"B{this_last - 5}").value = int(ME_holiday + 0.5)
        ws1.range(f"B{this_last - 4}").value = int(ME_sick + 0.5)
        ws1.range(f"B{this_last - 3}").value = int(ME_funeral + 0.5)
        ws1.range(f"B{this_last - 2}").value = int(ME_other + 0.5)
        ws1.range(f"B{this_last - 1}").value = int(ME_service + 0.5)
        ws1.range(f"B{this_last}").value = int(ME_sum + 0.5)
        ws1.range(f"C{this_last - 11}").value = int(EE_C + 0.5)
        ws1.range(f"C{this_last - 10}").value = int(EE_R + 0.5)
        ws1.range(f"C{this_last - 9}").value = int(EE_U + 0.5)
        ws1.range(f"C{this_last - 8}").value = int(EE_D + 0.5)
        ws1.range(f"C{this_last - 7}").value = int(EE_admin + 0.5)
        ws1.range(f"C{this_last - 6}").value = int(EE_vacation + 0.5)
        ws1.range(f"C{this_last - 5}").value = int(EE_holiday + 0.5)
        ws1.range(f"C{this_last - 4}").value = int(EE_sick + 0.5)
        ws1.range(f"C{this_last - 3}").value = int(EE_funeral + 0.5)
        ws1.range(f"C{this_last - 2}").value = int(EE_other + 0.5)
        ws1.range(f"C{this_last - 1}").value = int(EE_service + 0.5)
        ws1.range(f"C{this_last}").value = int(EE_sum + 0.5)
        try:
            wb.save(ot_hours)
        except Exception:
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {ot_hours} to allow editing", "Info", 0)
            return

        self.progress.emit(count + 1)
        sum_hours = data_folder + f"ENGINEERING TIMESHEET SUMMARY WITH BREAKOUT WE {last_sat}.xlsm"
        try:                                # Make excel sheet ready to edit (Summary)
            wb1 = xw.Book(sum_hours)
        except Exception as e:  # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, "Setup file ENGINEERING TIMESHEET SUMMARY WITH BREAKOUT WE "
                                                f"{last_sat} first", "Info", 0)
            print('Catch: ', e.__class__)
            return 1
        ws1 = wb1.sheets['Summary']
        ws1.range("B1").value = f"{last_sat_f}"  # set summary tab for new week
        ws1.range("B3").value = f"{last_sat_f}"
        ws1.charts[0].name = 'Chart 1'
        ws1.charts['Chart 1'].api[1].ChartTitle.Text = f"Engineering Overtime - {last_sat_f2}"
        #wb.sheets['Summary']["B1"].value = f"{next_sat_f}"  # set summary tab for new week
        #wb.sheets['Summary']["B3"].value = f"{next_sat_f}"
        #wb.sheets['Summary'].charts[0].name = 'Chart 1'
        #wb.sheets['Summary'].charts['Chart 1'].api[1].ChartTitle.Text = f"Engineering Overtime - {next_sat_f2}"

        last_no_ovt = 40
        this_last = last_no_ovt + num_ovt
        max_row = ws1.range('A' + str(ws1.cells.last_cell.row)).end('up').row
        move = this_last - max_row

        # Shift the data in the file depending on number of overtimes
        if move != 0:
            ws1.charts[0].name = 'Chart 1'
            wb1.sheets['Summary'].range(f"A{num_ovt + 6 - move}:F{max_row}").copy(wb1.sheets['Summary'].range(
                f"A{num_ovt + 6}:F{this_last}"))
            ws1.charts["Chart 1"].delete()  # Fix issues from moving a chart
            ws1.charts[0].name = 'Chart 1'

        if move < 0:  # Clear unneccessary data from moving content
            ws1.range(f"A{this_last + 1}:F{max_row}").clear()  # for shrink

        if num_ovt > 0:
            for border_id in range(7, 13):  # Reformat the table containing overtime employees
                ws1.range(f"A6:D{num_ovt + 5}").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A6:D{num_ovt + 5}").api.Borders(border_id).Weight = 2
        else:
            for border_id in range(7, 13):
                ws1.range(f"A5:D5").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A5:D5").api.Borders(border_id).Weight = 2

        for i in range(num_ovt):  # Populate the overtime table at the top
            ws1.cells(i + 6, 1).value = ovt[i][0]
            ws1.cells(i + 6, 2).value = int(ovt[i][1] + 0.5)
            ws1.cells(i + 6, 3).value = int(ovt[i][2] + 0.5)
            ws1.cells(i + 6, 4).value = int(ovt[i][3] + 0.5)

        # Populate new data for current week
        ws1.range("B4").value = int(ME_total + 0.5)
        ws1.range("C4").value = int(ME_ovt + 0.5)
        ws1.range("D4").value = int(ME_ovt_paid + 0.5)
        ws1.range("B5").value = int(EE_total + 0.5)
        ws1.range("C5").value = int(EE_ovt + 0.5)
        ws1.range("D5").value = int(EE_ovt_paid + 0.5)

        ws1.range(f"B{this_last - 11}").value = int(ME_C + 0.5)
        ws1.range(f"B{this_last - 10}").value = int(ME_R + 0.5)
        ws1.range(f"B{this_last - 9}").value = int(ME_U + 0.5)
        ws1.range(f"B{this_last - 8}").value = int(ME_D + 0.5)
        ws1.range(f"B{this_last - 7}").value = int(ME_admin + 0.5)
        ws1.range(f"B{this_last - 6}").value = int(ME_vacation + 0.5)
        ws1.range(f"B{this_last - 5}").value = int(ME_holiday + 0.5)
        ws1.range(f"B{this_last - 4}").value = int(ME_sick + 0.5)
        ws1.range(f"B{this_last - 3}").value = int(ME_funeral + 0.5)
        ws1.range(f"B{this_last - 2}").value = int(ME_other + 0.5)
        ws1.range(f"B{this_last - 1}").value = int(ME_service + 0.5)
        ws1.range(f"B{this_last}").value = int(ME_sum + 0.5)
        ws1.range(f"C{this_last - 11}").value = int(EE_C + 0.5)
        ws1.range(f"C{this_last - 10}").value = int(EE_R + 0.5)
        ws1.range(f"C{this_last - 9}").value = int(EE_U + 0.5)
        ws1.range(f"C{this_last - 8}").value = int(EE_D + 0.5)
        ws1.range(f"C{this_last - 7}").value = int(EE_admin + 0.5)
        ws1.range(f"C{this_last - 6}").value = int(EE_vacation + 0.5)
        ws1.range(f"C{this_last - 5}").value = int(EE_holiday + 0.5)
        ws1.range(f"C{this_last - 4}").value = int(EE_sick + 0.5)
        ws1.range(f"C{this_last - 3}").value = int(EE_funeral + 0.5)
        ws1.range(f"C{this_last - 2}").value = int(EE_other + 0.5)
        ws1.range(f"C{this_last - 1}").value = int(EE_service + 0.5)
        ws1.range(f"C{this_last}").value = int(EE_sum + 0.5)
        wb1.save()
        excel_app.kill()
        self.progress.emit(count + 2)

        self.SumTabCov()

    def SumTabCov(self):        # populate department tabs in timesheet summary
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        last_sat_f = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%Y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/WE {last_sat}/"
        #data_folder = f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/" \
        #              f"{cur_year} COV Timesheets/WE {last_sat}/"

        target_files = []
        for root, dirs, files in os.walk(data_folder):  # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))

        absences = []
        def get_tab(file_name):     # get summary tab based on file name
            return {
                data_folder + f'EE_AFTERMARKET WE {last_sat}.xlsm': 'EE_Aftermarket',
                data_folder + f'ee_aftermarket we {last_sat}.xlsm': 'EE_Aftermarket',
                data_folder + f'EE_CONTRACTOR WE {last_sat}.xlsm': 'EE_ConCoop',
                data_folder + f'ee_contractor we {last_sat}.xlsm': 'EE_ConCoop',
                data_folder + f'EE_COOP WE {last_sat}.xlsm': 'EE_ConCoop',
                data_folder + f'ee_coop we {last_sat}.xlsm': 'EE_ConCoop',
                data_folder + f'EE_CUSTCTRLS WE {last_sat}.xlsm': 'EE_CustCtrls',
                data_folder + f'ee_custctrls we {last_sat}.xlsm': 'EE_CustCtrls',
                data_folder + f'EE_INNOVATION WE {last_sat}.xlsm': 'EE_Innovation',
                data_folder + f'ee_innovation we {last_sat}.xlsm': 'EE_Innovation',
                data_folder + f'EE_PRODLINECTRLS WE {last_sat}.xlsm': 'EE_ProdLineCtrls',
                data_folder + f'ee_prodlinectrls we {last_sat}.xlsm': 'EE_ProdLineCtrls',
                data_folder + f'EE_SPECIALISTS WE {last_sat}.xlsm': 'EE_Specialists',
                data_folder + f'ee_specialists we {last_sat}.xlsm': 'EE_Specialists',
                data_folder + f'ME_APPS WE {last_sat}.xlsm': 'ME_Apps',
                data_folder + f'me_apps we {last_sat}.xlsm': 'ME_Apps',
                data_folder + f'ME_CONTRACTOR WE {last_sat}.xlsm': 'ME_ConCoop',
                data_folder + f'me_contractor we {last_sat}.xlsm': 'ME_ConCoop',
                data_folder + f'ME_COOP WE {last_sat}.xlsm': 'ME_ConCoop',
                data_folder + f'me_coop we {last_sat}.xlsm': 'ME_ConCoop',
                data_folder + f'ME_CUSTSOL WE {last_sat}.xlsm': 'ME_CustSol',
                data_folder + f'me_custsol we {last_sat}.xlsm': 'ME_CustSol',
                data_folder + f'ME_INNOVATION WE {last_sat}.xlsm': 'ME_Innovation',
                data_folder + f'me_innovation we {last_sat}.xlsm': 'ME_Innovation',
                data_folder + f'ME_PRODLINE WE {last_sat}.xlsm': 'ME_ProdLine',
                data_folder + f'me_prodline we {last_sat}.xlsm': 'ME_ProdLine',
                data_folder + f'ME_SYSENG WE {last_sat}.xlsm': 'ME_SysEng',
                data_folder + f'me_syseng we {last_sat}.xlsm': 'ME_SysEng'
            }.get(file_name, 'ME_ConCoop')

        def get_reason(x):      # get absence type based on row num
            return {
                6: ' hours vacation',
                7: ' hours holiday',
                8: ' hours sick',
                9: ' hours funeral',
                10: ' hours excuse'
            }.get(x, ' hours excuse')

        count = len(target_files) + 2
        for file in target_files:
            #print(file)
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if sheet != "Setup":
                    ws = wb[sheet]
                    #print(sheet)
                    tab, reason, week = "", "", ["", "", "", "", ""]
                    total, ovt, ovt_p, this, hours = "", "", "", "", 0
                    tab = get_tab(file)     # look into dictionary for file name and summary tab connection
                    name = ws["C3"].value
                    for i in range(4, 9):   # loop trough D6-H10
                        for j in range(6, 11):
                            if ws.cell(j, i).value is not None:
                                this = get_reason(j)
                                if week[i - 4] == "":
                                    week[i - 4] = f"{ws.cell(j, i).value}/{ws.cell(j, 1).value[0]}"
                                else:
                                    week[i - 4] = f"{week[i - 4]}\n{ws.cell(j, i).value}/{ws.cell(j, 1).value[0]}"
                                if reason == "":
                                    reason = str(ws.cell(j, i).value) + get_reason(j)
                                    hours = ws.cell(j, i).value
                                elif this in reason and '\n' not in reason:
                                    hours += ws.cell(j, i).value
                                    reason = str(hours) + get_reason(j)
                                elif this in reason:
                                    section = reason.split('\n')
                                    temp = get_reason(j)
                                    for k in range(len(section)):
                                        if temp in section[k]:
                                            time, part = section[k].split(" ", 1)
                                            time = float(time) + float(ws.cell(j, i).value)
                                            section[k] = str(time) + ' ' + part
                                        if k == 0:
                                            reason = section[k]
                                        elif k != 0:
                                            reason = reason + '\n' + section[k]
                                else:
                                    reason = reason + '\n' + str(ws.cell(j, i).value) + get_reason(j)
                    total = ws["J41"].value
                    ovt = ws["J42"].value
                    ovt_p = ws["J43"].value
                    absences.append((tab, name, week[0], week[1], week[2], week[3], week[4], reason, total, ovt, ovt_p))
            count += 1
            self.progress.emit(count)

        #print(absences)
        wb.close()

        absences.sort(key=lambda x: x[0])
        depts = {}
        all = 0
        for i in range(len(absences)):          # get number of employees per department
            if not any(absences[i][0] in sublist for sublist in depts):
                nums = sum(row.count(absences[i][0]) for row in absences)
                depts[absences[i][0]] = nums
                all += nums
        #print(depts)
        #print(all)

        sum_hours = data_folder / Path(f"ENGINEERING TIMESHEET SUMMARY WITH BREAKOUT WE {last_sat}.xlsm")
        #print(sum_hours)
        dept_key = list(depts.keys())
        idx = 0
        excel_app = xw.App(visible=False)
        wb = xw.Book(sum_hours)
        for i in range(len(depts)):         # use absences list to fill out timesheet summary tabs
            ws = wb.sheets[dept_key[i]]
            for j in range(depts[dept_key[i]]):
                ws.range(f"A{j + 7}").value = absences[idx][1]
                ws.range(f"B{j + 7}").value = absences[idx][2]
                ws.range(f"C{j + 7}").value = absences[idx][3]
                ws.range(f"D{j + 7}").value = absences[idx][4]
                ws.range(f"E{j + 7}").value = absences[idx][5]
                ws.range(f"F{j + 7}").value = absences[idx][6]
                ws.range(f"G{j + 7}").value = absences[idx][7]
                ws.range(f"H{j + 7}").value = absences[idx][8]
                ws.range(f"I{j + 7}").value = absences[idx][9]
                ws.range(f"J{j + 7}").value = absences[idx][10]
                idx += 1
            ws.range("B3").value = last_sat_f
            ws.range("I2").value = f'=SUM(H7:H{depts[dept_key[i]] + 6})'
            ws.range("I3").value = f'=SUM(I7:I{depts[dept_key[i]] + 6})'
            ws.range("I4").value = f'=SUM(J7:J{depts[dept_key[i]] + 6})'
            if ws.range(f'A{depts[dept_key[i]] + 7}').value is not None:        # clear extras not entered
                max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
                ws.range(f"A{depts[dept_key[i]] + 7}:J{max_row}").clear_contents()

        wb.save()
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"WE {last_sat} and Overtime COV {cur_year} complete", "Info", 0)

    def OvtHrsDav(self):
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        last_sat_f = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m-%d-%y')
        last_sat_f1 = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%Y')
        last_sat_f2 = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {last_sat}/"
        #data_folder = f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/" \
        #              f"{cur_year} DAV Timesheets/WE {last_sat}/"
        target_files = []
        for root, dirs, files in os.walk(data_folder):          # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))
        if target_files == []:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {data_folder}", "Info", 0)
            return
        ot_hours = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/Overtime {cur_year}.xlsx"
        #ot_hours = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/Overtime Hours/"
        #                f"overtime {cur_year}.xlsx")

        ovt = []
        EE_total, EE_ovt, EE_ovt_paid, ME_total, ME_ovt, ME_ovt_paid = 0, 0, 0, 0, 0, 0
        EE_C, EE_R, EE_U, EE_D, ME_C, ME_R, ME_U, ME_D = 0, 0, 0, 0, 0, 0, 0, 0
        EE_admin, EE_vacation, EE_holiday, EE_sick, EE_funeral, EE_other, EE_service = 0, 0, 0, 0, 0, 0, 0
        ME_admin, ME_vacation, ME_holiday, ME_sick, ME_funeral, ME_other, ME_service = 0, 0, 0, 0, 0, 0, 0

        count = 0
        self.maximum.emit((2 * len(target_files)) + 2)
        for file in target_files:
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if "EE" in file and sheet != "Setup":  # Obtain EE info
                    ws = wb[sheet]
                    #print(sheet)
                    EE_total += ws["J41"].value
                    EE_vacation += ws["J6"].value
                    EE_holiday += ws["J7"].value
                    EE_sick += ws["J8"].value
                    EE_funeral += ws["J9"].value
                    EE_other += ws["J10"].value
                    EE_admin += ws["J11"].value
                    if ws["J42"].value != 0:
                        if "EE_COOP" in file or "ee_coop" in file:
                            type = "EE Coop"
                        elif "EE_CONTRACTOR" in file or "ee_contractor" in file:
                            type = "EE Contractor"
                        else:
                            type = "Electrical"
                        EE_ovt += ws["J42"].value
                        EE_ovt_paid += ws["J43"].value
                        ovt_list = [ws["C3"].value, ws["J41"].value, ws["J42"].value, ws["J43"].value, type]
                        ovt.append(ovt_list)
                    for s in range(4):
                        if ws[f"J{14 + s}"].value != 0:
                            EE_service += ws[f"J{14 + s}"].value
                    for i in range(21):
                        if ws[f"J{20 + i}"].value != 0:
                            char = (ws[f"A{20 + i}"].value[0]).upper()
                            if char == "C":
                                EE_C += ws[f"J{20 + i}"].value
                            elif char == "R":
                                EE_R += ws[f"J{20 + i}"].value
                            elif char == "U":
                                EE_U += ws[f"J{20 + i}"].value
                            elif char == "D":
                                EE_D += ws[f"J{20 + i}"].value
                    #print(EE_C,EE_R,EE_U,EE_D,EE_admin,EE_vacation,EE_holiday,EE_sick,EE_funeral,EE_other,EE_service,EE_total)
                elif "ME" in file and sheet != "Setup":  # Obtain ME info
                    ws = wb[sheet]
                    #print(sheet)
                    ME_total += ws[f"J41"].value
                    ME_vacation += ws["J6"].value
                    ME_holiday += ws["J7"].value
                    ME_sick += ws["J8"].value
                    ME_funeral += ws["J9"].value
                    ME_other += ws["J10"].value
                    ME_admin += ws["J11"].value
                    if ws[f"J42"].value != 0:
                        if "ME_COOP" in file or "me_coop" in file:
                            type = "ME Coop"
                        elif "ME_CONTRACTOR" in file or "me_contractor" in file:
                            type = "ME Contractor"
                        else:
                            type = "Mechanical"
                        ME_ovt += ws[f"J42"].value
                        ME_ovt_paid += ws[f"J43"].value
                        ovt_list = [ws["C3"].value, ws[f"J41"].value, ws[f"J42"].value, ws[f"J43"].value, type]
                        ovt.append(ovt_list)
                    for s in range(4):
                        if ws[f"J{14 + s}"].value != 0:
                            ME_service += ws[f"J{14 + s}"].value
                    for i in range(21):
                        if ws[f"J{20 + i}"].value != 0:
                            char = (ws[f"A{20 + i}"].value[0]).upper()
                            if char == "C":
                                ME_C += ws[f"J{20 + i}"].value
                            elif char == "R":
                                ME_R += ws[f"J{20 + i}"].value
                            elif char == "U":
                                ME_U += ws[f"J{20 + i}"].value
                            elif char == "D":
                                ME_D += ws[f"J{20 + i}"].value
                    #print(ME_C,ME_R,ME_U,ME_D,ME_admin,ME_vacation,ME_holiday,ME_sick,ME_funeral,ME_other,ME_service,ME_total)
            count += 1
            self.progress.emit(count)

        EE_sum = EE_C+EE_R+EE_U+EE_D+EE_admin+EE_vacation+EE_holiday+EE_sick+EE_funeral+EE_other+EE_service
        ME_sum = ME_C+ME_R+ME_U+ME_D+ME_admin+ME_vacation+ME_holiday+ME_sick+ME_funeral+ME_other+ME_service

        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        try:
            wb = xw.Book(ot_hours)                                      # Make excel sheet ready to edit (Yearly Overtime)
        except Exception:              # make sure yearly overtime file exist
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Couldn't locate file {ot_hours}", "Info", 0)
            return
        try:
            ws1 = wb.sheets[f'{last_sat_f}']
        except Exception as e:  # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Setup excel sheet {last_sat_f} first", "Info", 0)
            print('Catch: ', e.__class__)
            return 1
        last_no_ovt = 39
        num_ovt = len(ovt)
        this_last = last_no_ovt + num_ovt
        max_row = ws1.range('A' + str(ws1.cells.last_cell.row)).end('up').row
        move = this_last - max_row

        # Shift the data in the file depending on number of overtimes
        if move != 0:
            wb.sheets[f'{last_sat_f}'].range(f"A{num_ovt + 3 - move}:F{max_row}").copy(wb.sheets[f'{last_sat_f}']
                                                                                .range(f"A{num_ovt + 3}:F{this_last}"))
            ws1.charts["Chart 1"].delete()  # Fix issues from moving a chart
            ws1.charts[0].name = "Chart 1"
            ws1.charts["Chart 1"].set_source_data(ws1.range(f"B{num_ovt + 4}:E{num_ovt + 5}"))
        if move < 0:  # Clear unneccessary data from moving content
            ws1.range(f"A{this_last + 1}:F{max_row}").clear()  # for shrink
        elif move > 0:
            ws1.range(f"B{num_ovt - move + 3}:E{num_ovt - move + 4}").clear()  # for grow

        if num_ovt > 0:
            for border_id in range(7, 13):  # Reformat the table containing overtime employees
                ws1.range(f"A3:E{num_ovt + 2}").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A3:E{num_ovt + 2}").api.Borders(border_id).Weight = 2
        if move > 0:
            ws1.range(f"A3:D{num_ovt + 2}").api.Font.Size = 10
            ws1.range(f"A3:D{num_ovt + 2}").api.Font.Name = "Arial"

        for i in range(num_ovt):  # Populate the overtime table at the top
            ws1.cells(i + 3, 1).value = ovt[i][0]
            ws1.cells(i + 3, 2).value = int(ovt[i][1] + 0.5)
            ws1.cells(i + 3, 3).value = int(ovt[i][2] + 0.5)
            ws1.cells(i + 3, 4).value = int(ovt[i][3] + 0.5)
            ws1.cells(i + 3, 5).value = ovt[i][4]

        # Populate new data for current week
        ws1.range(f"B{num_ovt + 5}").value = int(ME_total + 0.5)
        ws1.range(f"C{num_ovt + 5}").value = int(ME_ovt + 0.5)
        ws1.range(f"D{num_ovt + 5}").value = int(EE_total + 0.5)
        ws1.range(f"E{num_ovt + 5}").value = int(EE_ovt + 0.5)

        ws1.range(f"B{this_last - 11}").value = int(ME_C + 0.5)
        ws1.range(f"B{this_last - 10}").value = int(ME_R + 0.5)
        ws1.range(f"B{this_last - 9}").value = int(ME_U + 0.5)
        ws1.range(f"B{this_last - 8}").value = int(ME_D + 0.5)
        ws1.range(f"B{this_last - 7}").value = int(ME_admin + 0.5)
        ws1.range(f"B{this_last - 6}").value = int(ME_vacation + 0.5)
        ws1.range(f"B{this_last - 5}").value = int(ME_holiday + 0.5)
        ws1.range(f"B{this_last - 4}").value = int(ME_sick + 0.5)
        ws1.range(f"B{this_last - 3}").value = int(ME_funeral + 0.5)
        ws1.range(f"B{this_last - 2}").value = int(ME_other + 0.5)
        ws1.range(f"B{this_last - 1}").value = int(ME_service + 0.5)
        ws1.range(f"B{this_last}").value = int(ME_sum + 0.5)
        ws1.range(f"C{this_last - 11}").value = int(EE_C + 0.5)
        ws1.range(f"C{this_last - 10}").value = int(EE_R + 0.5)
        ws1.range(f"C{this_last - 9}").value = int(EE_U + 0.5)
        ws1.range(f"C{this_last - 8}").value = int(EE_D + 0.5)
        ws1.range(f"C{this_last - 7}").value = int(EE_admin + 0.5)
        ws1.range(f"C{this_last - 6}").value = int(EE_vacation + 0.5)
        ws1.range(f"C{this_last - 5}").value = int(EE_holiday + 0.5)
        ws1.range(f"C{this_last - 4}").value = int(EE_sick + 0.5)
        ws1.range(f"C{this_last - 3}").value = int(EE_funeral + 0.5)
        ws1.range(f"C{this_last - 2}").value = int(EE_other + 0.5)
        ws1.range(f"C{this_last - 1}").value = int(EE_service + 0.5)
        ws1.range(f"C{this_last}").value = int(EE_sum + 0.5)
        try:
            wb.save(ot_hours)
        except Exception:
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {ot_hours} to allow editing", "Info", 0)
            return

        self.progress.emit(count + 1)
        last_sat_s = last_sat.replace(' ', '_')
        sum_hours = data_folder + f"ENGINEERING TIMESHEET SUMMARY WE {last_sat_s}_WithBreakout.xlsm"
        #print(sum_hours)
        try:  # Make excel sheet ready to edit (Summary)
            wb1 = xw.Book(sum_hours)
        except Exception as e:  # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Setup file ENGINEERING TIMESHEET SUMMARY WE {last_sat_s}"
                                                f"_WithBreakout first", "Info", 0)
            print('Catch: ', e.__class__)
            return 1
        ws1 = wb1.sheets['Summary']
        ws1.range("B1").value = f"{last_sat_f1}"  # set summary tab for new week
        ws1.range("B3").value = f"{last_sat_f1}"
        ws1.charts[0].name = 'Chart 1'
        ws1.charts['Chart 1'].api[1].ChartTitle.Text = f"Engineering Overtime - {last_sat_f2}"

        last_no_ovt = 40
        this_last = last_no_ovt + num_ovt
        max_row = ws1.range('A' + str(ws1.cells.last_cell.row)).end('up').row
        move = this_last - max_row

        # Shift the data in the file depending on number of overtimes
        if move != 0:
            ws1.charts[0].name = 'Chart 1'
            wb1.sheets['Summary'].range(f"A{num_ovt + 6 - move}:F{max_row}").copy(wb1.sheets['Summary'].range(
                f"A{num_ovt + 6}:F{this_last}"))
            ws1.charts["Chart 1"].delete()  # Fix issues from moving a chart
            ws1.charts[0].name = 'Chart 1'

        if move < 0:  # Clear unneccessary data from moving content
            ws1.range(f"A{this_last + 1}:F{max_row}").clear()  # for shrink

        if num_ovt > 0:
            for border_id in range(7, 13):  # Reformat the table containing overtime employees
                ws1.range(f"A6:D{num_ovt + 5}").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A6:D{num_ovt + 5}").api.Borders(border_id).Weight = 2
        else:
            for border_id in range(7, 13):
                ws1.range(f"A5:D5").api.Borders(border_id).LineStyle = 1
                ws1.range(f"A5:D5").api.Borders(border_id).Weight = 2

        for i in range(num_ovt):  # Populate the overtime table at the top
            ws1.cells(i + 6, 1).value = ovt[i][0]
            ws1.cells(i + 6, 2).value = int(ovt[i][1] + 0.5)
            ws1.cells(i + 6, 3).value = int(ovt[i][2] + 0.5)
            ws1.cells(i + 6, 4).value = int(ovt[i][3] + 0.5)

        # Populate new data for current week
        ws1.range("B4").value = int(ME_total + 0.5)
        ws1.range("C4").value = int(ME_ovt + 0.5)
        ws1.range("D4").value = int(ME_ovt_paid + 0.5)
        ws1.range("B5").value = int(EE_total + 0.5)
        ws1.range("C5").value = int(EE_ovt + 0.5)
        ws1.range("D5").value = int(EE_ovt_paid + 0.5)

        ws1.range(f"B{this_last - 11}").value = int(ME_C + 0.5)
        ws1.range(f"B{this_last - 10}").value = int(ME_R + 0.5)
        ws1.range(f"B{this_last - 9}").value = int(ME_U + 0.5)
        ws1.range(f"B{this_last - 8}").value = int(ME_D + 0.5)
        ws1.range(f"B{this_last - 7}").value = int(ME_admin + 0.5)
        ws1.range(f"B{this_last - 6}").value = int(ME_vacation + 0.5)
        ws1.range(f"B{this_last - 5}").value = int(ME_holiday + 0.5)
        ws1.range(f"B{this_last - 4}").value = int(ME_sick + 0.5)
        ws1.range(f"B{this_last - 3}").value = int(ME_funeral + 0.5)
        ws1.range(f"B{this_last - 2}").value = int(ME_other + 0.5)
        ws1.range(f"B{this_last - 1}").value = int(ME_service + 0.5)
        ws1.range(f"B{this_last}").value = int(ME_sum + 0.5)
        ws1.range(f"C{this_last - 11}").value = int(EE_C + 0.5)
        ws1.range(f"C{this_last - 10}").value = int(EE_R + 0.5)
        ws1.range(f"C{this_last - 9}").value = int(EE_U + 0.5)
        ws1.range(f"C{this_last - 8}").value = int(EE_D + 0.5)
        ws1.range(f"C{this_last - 7}").value = int(EE_admin + 0.5)
        ws1.range(f"C{this_last - 6}").value = int(EE_vacation + 0.5)
        ws1.range(f"C{this_last - 5}").value = int(EE_holiday + 0.5)
        ws1.range(f"C{this_last - 4}").value = int(EE_sick + 0.5)
        ws1.range(f"C{this_last - 3}").value = int(EE_funeral + 0.5)
        ws1.range(f"C{this_last - 2}").value = int(EE_other + 0.5)
        ws1.range(f"C{this_last - 1}").value = int(EE_service + 0.5)
        ws1.range(f"C{this_last}").value = int(EE_sum + 0.5)
        wb1.save()
        wb1.close()
        excel_app.kill()
        self.progress.emit(count + 2)

        self.SumTabDav()

    def SumTabDav(self):            # populate department tabs in timesheet summary
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        last_sat_f = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m/%d/%Y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {last_sat}/"
        #data_folder = f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/" \
        #              f"{cur_year} DAV Timesheets/WE {last_sat}/"
        target_files = []
        for root, dirs, files in os.walk(data_folder):  # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))
        absences = []

        def get_tab(file_name):  # get summary tab based on file name
            return {
                data_folder + f'EE_CustomCtrls WE {last_sat}.xlsm': 'EE_CustomCtrls',
                data_folder + f'ee_customctrls we {last_sat}.xlsm': 'EE_CustomCtrls',
                data_folder + f'EE_Innovation WE {last_sat}.xlsm': 'EE_Innovation',
                data_folder + f'ee_innovation we {last_sat}.xlsm': 'EE_Innovation',
                data_folder + f'EE_ProdLineCtrls WE {last_sat}.xlsm': 'EE_ProdLineCtrls',
                data_folder + f'ee_prodlinectrls we {last_sat}.xlsm': 'EE_ProdLineCtrls',
                data_folder + f'EE_Specialists WE {last_sat}.xlsm': 'EE_Specialists',
                data_folder + f'ee_specialists we {last_sat}.xlsm': 'EE_Specialists',
                data_folder + f'ME_Apps WE {last_sat}.xlsm': 'ME_Apps',
                data_folder + f'me_apps we {last_sat}.xlsm': 'ME_Apps',
                data_folder + f'ME_CustomSol WE {last_sat}.xlsm': 'ME_CustomSol',
                data_folder + f'me_customsol we {last_sat}.xlsm': 'ME_CustomSol',
                data_folder + f'ME_Innovation WE {last_sat}.xlsm': 'ME_Innovation',
                data_folder + f'me_innovation we {last_sat}.xlsm': 'ME_Innovation',
                data_folder + f'ME_ProdLine WE {last_sat}.xlsm': 'ME_ProdLine',
                data_folder + f'me_prodline we {last_sat}.xlsm': 'ME_ProdLine',
                data_folder + f'ME_Systems WE {last_sat}.xlsm': 'ME_Systems',
                data_folder + f'me_systems we {last_sat}.xlsm': 'ME_Systems'
            }.get(file_name, 'ME_Systems')

        def get_reason(x):  # get absence type based on row num
            return {
                6: ' hours vacation',
                7: ' hours holiday',
                8: ' hours sick',
                9: ' hours funeral',
                10: ' hours excuse'
            }.get(x, ' hours excuse')

        count = len(target_files) + 2
        for file in target_files:
            #print(file)
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if sheet != "Setup":
                    ws = wb[sheet]
                    #print(sheet)
                    tab, reason, week = "", "", ["", "", "", "", ""]
                    total, ovt, ovt_p, this, hours = "", "", "", "", 0
                    tab = get_tab(file)  # look into dictionary for file name and summary tab connection
                    name = ws["C3"].value
                    for i in range(4, 9):  # loop trough D6-H10
                        for j in range(6, 11):
                            if ws.cell(j, i).value is not None:
                                this = get_reason(j)
                                if week[i - 4] == "":
                                    week[i - 4] = f"{ws.cell(j, i).value}/{ws.cell(j, 1).value[0]}"
                                else:
                                    week[i - 4] = f"{week[i - 4]}\n{ws.cell(j, i).value}/{ws.cell(j, 1).value[0]}"
                                if reason == "":
                                    reason = str(ws.cell(j, i).value) + get_reason(j)
                                    hours = ws.cell(j, i).value
                                elif this in reason and '\n' not in reason:
                                    hours += ws.cell(j, i).value
                                    reason = str(hours) + get_reason(j)
                                elif this in reason:
                                    section = reason.split('\n')
                                    temp = get_reason(j)
                                    for k in range(len(section)):
                                        if temp in section[k]:
                                            time, part = section[k].split(" ", 1)
                                            time = float(time) + float(ws.cell(j, i).value)
                                            section[k] = str(time) + ' ' + part
                                        if k == 0:
                                            reason = section[k]
                                        elif k != 0:
                                            reason = reason + '\n' + section[k]
                                else:
                                    reason = reason + '\n' + str(ws.cell(j, i).value) + get_reason(j)
                    total = ws["J41"].value
                    ovt = ws["J42"].value
                    ovt_p = ws["J43"].value
                    absences.append((tab, name, week[0], week[1], week[2], week[3], week[4], reason, total, ovt, ovt_p))
            count += 1
            self.progress.emit(count)

        #print(absences)
        wb.close()

        absences.sort(key=lambda x: x[0])
        depts = {}
        all = 0
        for i in range(len(absences)):          # get number of employees per department
            if not any(absences[i][0] in sublist for sublist in depts):
                nums = sum(row.count(absences[i][0]) for row in absences)
                depts[absences[i][0]] = nums
                all += nums
        #print(depts)
        #print(all)

        last_sat_s = last_sat.replace(' ', '_')
        sum_hours = data_folder / Path(f"ENGINEERING TIMESHEET SUMMARY WE {last_sat_s}_WithBreakout.xlsm")
        #print(sum_hours)

        dept_key = list(depts.keys())
        idx = 0
        excel_app = xw.App(visible=False)
        wb = xw.Book(sum_hours)
        for i in range(len(depts)):         # use absences list to fill out timesheet summary tabs
            ws = wb.sheets[dept_key[i]]
            for j in range(depts[dept_key[i]]):
                ws.range(f"A{j + 7}").value = absences[idx][1]
                ws.range(f"B{j + 7}").value = absences[idx][2]
                ws.range(f"C{j + 7}").value = absences[idx][3]
                ws.range(f"D{j + 7}").value = absences[idx][4]
                ws.range(f"E{j + 7}").value = absences[idx][5]
                ws.range(f"F{j + 7}").value = absences[idx][6]
                ws.range(f"G{j + 7}").value = absences[idx][7]
                ws.range(f"H{j + 7}").value = absences[idx][8]
                ws.range(f"I{j + 7}").value = absences[idx][9]
                ws.range(f"J{j + 7}").value = absences[idx][10]
                idx += 1
            ws.range("B3").value = last_sat_f
            ws.range("I2").value = f'=SUM(H7:H{depts[dept_key[i]] + 6})'
            ws.range("I3").value = f'=SUM(I7:I{depts[dept_key[i]] + 6})'
            ws.range("I4").value = f'=SUM(J7:J{depts[dept_key[i]] + 6})'
            if ws.range(f'A{depts[dept_key[i]] + 7}').value is not None:  # clear extras not entered
                max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
                ws.range(f"A{depts[dept_key[i]] + 7}:J{max_row}").clear_contents()

        wb.save()
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"WE {last_sat} and Overtime {cur_year} complete", "Info", 0)

    def PopSapCov(self):
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/WE {last_sat}/"
        #data_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/"
        #                   f"{cur_year} COV Timesheets/WE {last_sat}/")
        target_files = []
        for root, dirs, files in os.walk(data_folder):  # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))
        count = 0
        self.maximum.emit(len(target_files) + 9)
        sap_proj = 21
        week_len = 7
        sap_entries = []
        last = [0, '', '', '']
        for file in target_files:
            wb = openpyxl.load_workbook(file, data_only=True) #, read_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if sheet != "Setup":                        # Obtain info unless setup tab
                    ws = wb[sheet]
                    #print(sheet)
                    for i in range(week_len):
                        for j in range(sap_proj):
                            hours = ws.cell(row=j + 20, column=i + 3).value
                            #print(hours)
                            if " " in str(hours):                   # removes blank spaces in response
                                hours = str(hours).replace(" ", "")
                            if hours is not None and hours != '':         # search file for hours data
                                confirm_num = ws.cell(row=j + 20, column=12).value
                                if confirm_num is None or len(confirm_num) < 5:
                                    self.finished.emit()
                                    pythoncom.CoUninitialize()
                                    ctypes.windll.user32.MessageBoxW(0, f"Confirmation number is missing for {sheet}",
                                                                     "Info", 0)
                                    return
                                try:
                                    hours = float(hours)
                                except Exception as e:  # Catch invalid values
                                    self.finished.emit()
                                    pythoncom.CoUninitialize()
                                    ctypes.windll.user32.MessageBoxW(0, "Invalid Value: Use \"Remove spaces\" button",
                                                                     "Info", 0)
                                    return
                                confirm, waste = confirm_num.split(' ')
                                unf_date = ws.cell(row=4, column=i + 3).value
                                mark_date = datetime.datetime.strptime(str(unf_date), '%Y-%m-%d %H:%M:%S').strftime('%m%d%Y')
                                if last[0] == ws['K3'].value and last[1] == mark_date:
                                    start = str(last[2])
                                    start_time = last[3]
                                else:
                                    start = "8:00:00"
                                    start_time = datetime.datetime.strptime(str(start), '%H:%M:%S').strftime('%I:%M:%S %p')
                                h, m, s = start.split(':')
                                #print(f'{h} {m} {s}')
                                hrs = str("{:.2f}".format(hours))
                                #print(hrs)
                                h2, m2 = hrs.split('.')
                                m2 = int(m2) / 100 * 60
                                #print(f'{h2} {m2}')
                                end = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s)) + \
                                      datetime.timedelta(hours=int(h2), minutes=int(m2))
                                if end.days >= 1:
                                    #end = end - datetime.timedelta(days=1)
                                    #ctypes.windll.user32.MessageBoxW(0, f"Hours for the day is or surpassed 16 for "
                                    #                                    f"{sheet}", "Info", 0)
                                    end_time = None
                                else:
                                    end_time = datetime.datetime.strptime(str(end), '%H:%M:%S').strftime('%I:%M:%S %p')
                                time_data = [ws['K3'].value, confirm, str(mark_date), hrs, start_time, end_time]
                                sap_entries.append(time_data)
                                last = [ws['K3'].value, mark_date, end, end_time]
            count += 1
            self.progress.emit(count)
        sap_entries = sorted(sap_entries, key=lambda x: x[0])
        #print(sap_entries)
        #print(len(sap_entries))

        sap_file = data_folder / Path(f"engineering timesheet summary with breakout we {last_sat}.xlsm")

        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        count += 1
        self.progress.emit(count)
        try:
            wb = xw.Book(sap_file)
        except Exception as e:  # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, "Setup file engineering timesheet summary with breakout we "
                                                f"{last_sat} first", "Info", 0)
            return 1
        ws = wb.sheets["SAP Hours"]

        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        #print(max_row)
        if max_row > 1:
            ws.range(f"A2:G{max_row}").clear_contents()

        xw.Range(f"A2:H{max_row}").api.Interior.ColorIndex = 0

        for k in range(len(sap_entries)):
            ws.range(f"A{k + 2}").value = sap_entries[k][0]
            ws.range(f"B{k + 2}").value = sap_entries[k][1]
            ws.range(f"C{k + 2}").value = sap_entries[k][2]
            ws.range(f"D{k + 2}").value = sap_entries[k][3]
            ws.range(f"E{k + 2}").value = sap_entries[k][4]
            ws.range(f"F{k + 2}").value = sap_entries[k][5]
            if ws.range(f"F{k + 2}").value == None:
                xw.Range(f"A{k + 2}:H{k + 2}").api.Interior.ColorIndex = 46     # orange highlight if >= 16 hours a day
            self.progress.emit(count + (k/(len(sap_entries)-1)*4))

        wb.save()
        excel_app.kill()

        self.confirmation_cov(sap_file, sap_entries, count)

    def confirmation_cov(self, file, entry, count):
        self.progress.emit(count + 5)
        data_folder = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/")
        fname = data_folder / "CN47N.XLSX"
        df = pd.read_excel(fname, sheet_name="Sheet1", header=[0])

        df.columns = ("Plant", "Network", "Activity", "Status", "Project definition", "WBS element", "Activity desc.",
                      "Work center", "Confirmation", "Actual duration", "Earliest start date (basic)",
                      "Actual finish date", "Work", "Actual work", "Processing % of work", "Remaining work")

        #print(df)

        self.progress.emit(count + 6)
        excel_app = xw.App(visible=False)
        self.progress.emit(count + 7)
        wb = xw.Book(file)
        ws = wb.sheets["SAP Hours"]
        ws.activate()

        confirm = df['Confirmation']
        status = df['Status']
        conf = []

        for j in range(len(confirm) - 1):               # obtain list of valid confirm nums and teco status
            con = str(confirm[j])
            con, junk = con.split('.')
            if "TECO" in str(status[j]):
                stat = 'teco'
            else:
                stat = 'active'
            conf.append((con, stat))

        for i in range(len(entry)):                         # highlights row confirmation
            IN_CN47N = False
            for j in range(len(conf)):
                if entry[i][1] == conf[j][0] and conf[j][1] == 'teco':          # blue if confirm num teco in cn47n
                    xw.Range(f"A{i + 2}:H{i + 2}").api.Interior.ColorIndex = 5
                    new_conf = self.change_confirm_num(j, conf)
                    ws.range(f"B{i + 2}").value = new_conf
                    IN_CN47N = True
                    break
                elif entry[i][1] == conf[j][0]:                             # no highlight if in cn47n and active
                    IN_CN47N = True
                    break
            if IN_CN47N == False:                   # yellow if confirm num not in cn47n
                xw.Range(f"A{i + 2}:H{i + 2}").api.Interior.ColorIndex = 6
        self.progress.emit(count + 8)

        wb.save()
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"SAP Hours in {file} complete", "Info", 0)

    def PopSapDav(self):
        today = datetime.date.today()  # Initialize date as last sat (for Ovt Hours)
        idx = (today.weekday() + 1) % 7
        sat = today - datetime.timedelta(idx + 1)
        last_sat = datetime.datetime.strptime(str(sat), '%Y-%m-%d').strftime('%m %d %y')
        cur_year = sat.year

        data_folder = f"C:/Users/Austin.Kidwell/Desktop/Over Time Hours/WE {last_sat}/"
        #data_folder = Path(f"//CPROME/Eng_Share/System Engineering/Visual Management/Timesheets/"
        #                   f"{cur_year} DAV Timesheets/WE {last_sat}/")
        target_files = []
        for root, dirs, files in os.walk(data_folder):  # Obtain files to retrieve data from
            for file in files:
                if last_sat in file:
                    if "SUMMARY" not in file and "summary" not in file:
                        ftarget = os.path.join(root, file)
                        target_files.append(ftarget)

        #print(len(target_files))
        count = 0
        self.maximum.emit(len(target_files) + 9)
        sap_proj = 21
        week_len = 7
        sap_entries = []
        last = [0, '', '', '']
        for file in target_files:
            wb = openpyxl.load_workbook(file, data_only=True)  # Make excel sheet ready to edit
            for sheet in wb.sheetnames:
                if sheet != "Setup":  # Obtain info unless setup tab
                    ws = wb[sheet]
                    #print(sheet)
                    for i in range(week_len):
                        for j in range(sap_proj):
                            hours = ws.cell(row=j + 20, column=i + 3).value
                            if " " in str(hours):                   # removes blank spaces in response
                                hours = str(hours).replace(" ", "")
                            if hours is not None and hours != '':   # search file for hours data
                                confirm_num = ws.cell(row=j + 20, column=12).value
                                if confirm_num is None or len(confirm_num) < 5:
                                    self.finished.emit()
                                    pythoncom.CoUninitialize()
                                    ctypes.windll.user32.MessageBoxW(0, f"Confirmation number is missing for {sheet}",
                                                                     "Info", 0)
                                    return
                                try:
                                    hours = float(hours)
                                except Exception as e:  # Catch invalid values
                                    self.finished.emit()
                                    pythoncom.CoUninitialize()
                                    ctypes.windll.user32.MessageBoxW(0, "Invalid Value: Use \"Remove spaces\" button",
                                                                     "Info", 0)
                                    return
                                confirm, waste = confirm_num.split(' ')
                                unf_date = ws.cell(row=4, column=i + 3).value
                                mark_date = datetime.datetime.strptime(str(unf_date), '%Y-%m-%d %H:%M:%S').strftime(
                                    '%m%d%Y')
                                if last[0] == ws['K3'].value and last[1] == mark_date:
                                    start = str(last[2])
                                    start_time = last[3]
                                else:
                                    start = "8:00:00"
                                    start_time = datetime.datetime.strptime(str(start), '%H:%M:%S').strftime(
                                        '%I:%M:%S %p')
                                h, m, s = start.split(':')
                                #print(f'{h} {m} {s}')
                                hrs = str("{:.2f}".format(hours))
                                #print(hrs)
                                h2, m2 = hrs.split('.')
                                m2 = int(m2) / 100 * 60
                                #print(f'{h2} {m2}')
                                end = datetime.timedelta(hours=int(h), minutes=int(m), seconds=int(s)) + \
                                      datetime.timedelta(hours=int(h2), minutes=int(m2))
                                if end.days >= 1:
                                    #ctypes.windll.user32.MessageBoxW(0, f"Hours for the day is or surpassed 23 for "
                                    #                                    f"{sheet}", "Info", 0)
                                    end_time = None
                                else:
                                    end_time = datetime.datetime.strptime(str(end), '%H:%M:%S').strftime('%I:%M:%S %p')
                                time_data = [ws['K3'].value, confirm, str(mark_date), hrs, start_time, end_time]
                                sap_entries.append(time_data)
                                last = [ws['K3'].value, mark_date, end, end_time]
            count += 1
            self.progress.emit(count)
        sap_entries = sorted(sap_entries, key=lambda x: x[0])
        #print(sap_entries)
        #print(len(sap_entries))

        last_sat_s = last_sat.replace(' ', '_')
        sap_file = data_folder / Path(f"ENGINEERING TIMESHEET SUMMARY WE {last_sat_s}_WithBreakout.xlsm")

        pythoncom.CoInitialize()
        excel_app = xw.App(visible=False)
        count += 1
        self.progress.emit(count)
        try:
            wb = xw.Book(sap_file)
        except Exception as e:  # must create sheet before populating it
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, "Setup file ENGINEERING TIMESHEET SUMMARY WITH BREAKOUT WE "
                                                f"{last_sat_s} first", "Info", 0)
            return 1
        ws = wb.sheets["SAP Hours"]

        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        #print(max_row)
        if max_row > 1:
            ws.range(f"A2:G{max_row}").clear_contents()

        xw.Range(f"A2:H{max_row}").api.Interior.ColorIndex = 0

        for k in range(len(sap_entries)):               # populate sap hours with timesheet info
            ws.range(f"A{k + 2}").value = sap_entries[k][0]
            ws.range(f"B{k + 2}").value = sap_entries[k][1]
            ws.range(f"C{k + 2}").value = sap_entries[k][2]
            ws.range(f"D{k + 2}").value = sap_entries[k][3]
            ws.range(f"E{k + 2}").value = sap_entries[k][4]
            ws.range(f"F{k + 2}").value = sap_entries[k][5]
            if ws.range(f"F{k + 2}").value == None:
                xw.Range(f"A{k + 2}:H{k + 2}").api.Interior.ColorIndex = 46     # orange highlight if >= 16 hours a day
            self.progress.emit(count + (k / (len(sap_entries) - 1) * 4))

        wb.save()
        excel_app.kill()

        self.confirmation_dav(sap_file, sap_entries, count)

    def confirmation_dav(self, file, entry, count):
        self.progress.emit(count + 5)
        data_folder = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/")
        fname = data_folder / "CN47N.XLSX"
        df = pd.read_excel(fname, sheet_name="Sheet1", header=[0])

        df.columns = ("Plant", "Network", "Activity", "Status", "Project definition", "WBS element", "Activity desc.",
                      "Work center", "Confirmation", "Actual duration", "Earliest start date (basic)",
                      "Actual finish date", "Work", "Actual work", "Processing % of work", "Remaining work")

        #print(df)

        self.progress.emit(count + 6)
        excel_app = xw.App(visible=False)
        self.progress.emit(count + 7)
        wb = xw.Book(file)
        ws = wb.sheets["SAP Hours"]
        ws.activate()

        confirm = df['Confirmation']
        status = df['Status']
        conf = []

        for j in range(len(confirm) - 1):  # obtain list of valid confirm nums and teco status
            con = str(confirm[j])
            con, junk = con.split('.')
            if "TECO" in str(status[j]):
                stat = 'teco'
            else:
                stat = 'active'
            conf.append((con, stat))

        for i in range(len(entry)):  # highlights row confirmation
            IN_CN47N = False
            for j in range(len(conf)):
                if entry[i][1] == conf[j][0] and conf[j][1] == 'teco':  # blue if confirm num teco in cn47n
                    xw.Range(f"A{i + 2}:H{i + 2}").api.Interior.ColorIndex = 5
                    new_conf = self.change_confirm_num(j, conf)
                    ws.range(f"B{i + 2}").value = new_conf
                    IN_CN47N = True
                    break
                elif entry[i][1] == conf[j][0]:  # no highlight if in cn47n and active
                    IN_CN47N = True
                    break
            if IN_CN47N == False:  # yellow if confirm num not in cn47n
                xw.Range(f"A{i + 2}:H{i + 2}").api.Interior.ColorIndex = 6
        self.progress.emit(count + 8)

        wb.save()
        excel_app.kill()
        self.progress.emit(0)
        self.finished.emit()
        pythoncom.CoUninitialize()
        ctypes.windll.user32.MessageBoxW(0, f"SAP Hours in {file} complete", "Info", 0)

    def change_confirm_num(self, idx, list):
        for i in range(idx, -1, -1):
            if list[i][1] == 'active':
                return list[i][0]
