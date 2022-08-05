
import os
import re
import ctypes
import pandas as pd
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
import datetime
from PyQt5.QtCore import QObject, pyqtSignal
import numpy as np
from pathlib import Path
import xlwings as xw
import pythoncom
from bisect import bisect


class PrjDbd(QObject):
    maximum = pyqtSignal(int)  # signals to communicate worker thread with main thread
    progress = pyqtSignal(int)
    finished = pyqtSignal()

    def EngH2ProjD(self):
        eng_hours = "C:/Users/Austin.Kidwell/Desktop/EngHrs.xlsx"
        #eng_hours = Path("//CPROME/Eng_Share/System Engineering/Visual Management/EngHours/EngHrs.xlsx")
        proj_dash = "C:/Users/Austin.Kidwell/Desktop/Project Dashboard R2.xlsx"
        #proj_dash = Path("//CPROME/Eng_Share/System Engineering/Visual Management/Test Folder/Project Dashboard R2.xlsx")

        try:
            df = pd.read_excel(eng_hours, sheet_name="Budget Hours", header=[4])
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {eng_hours}", "Info", 0)
            return

        #print(df)

        df_mask = df['Order'].str.contains('D009')
        df = df[~df_mask]
        cov_mask = df['COV/DAV'].str.contains('COV')
        cov_proj = df[cov_mask]
        dav_mask = df['COV/DAV'].str.contains('DAV')
        dav_proj = df[dav_mask]
        #print(len(cov_proj))
        #print(len(dav_proj))

        try:
            wb = openpyxl.load_workbook(proj_dash)
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Excel workbook {proj_dash} needs setup", "Info", 0)
            return
        ws = wb['Projects']

        last_row = 0
        for row in ws:
            if row[2].value != None:
                last_row += 1
        #print(last_row)
        cov_proj.index = range(len(cov_proj))  # Reset index
        dav_proj.index = range(len(dav_proj))

        i = 0
        self.maximum.emit(len(cov_proj) + len(dav_proj))
        # Update project hours with EngHrs info
        while ws[f"A{i + 2}"].value is not None:
            for j in range(len(cov_proj)):
                if ws[f"A{i + 2}"].value != 'COV' or ws[f"B{i + 2}"].value is None:
                    break
                elif (ws[f"B{i + 2}"].value).strip() == cov_proj['Order'][j]:
                    ws[f'H{i + 2}'] = int(cov_proj['As Sold\nBudget'][j] + 0.5)
                    ws[f'I{i + 2}'] = int(cov_proj['To Date\nActuals'][j] + 0.5)
                    ws[f'L{i + 2}'] = int(cov_proj['As Sold Budget'][j] + 0.5)
                    ws[f'M{i + 2}'] = int(cov_proj['To Date\nActuals.1'][j] + 0.5)
            for k in range(len(dav_proj)):
                if ws[f"A{i + 2}"].value != 'DAV' or ws[f"B{i + 2}"].value is None:
                    break
                elif (ws[f"B{i + 2}"].value).strip() == dav_proj['Order'][k]:
                    ws[f'H{i + 2}'] = int(dav_proj['As Sold\nBudget'][k] + 0.5)
                    ws[f'I{i + 2}'] = int(dav_proj['To Date\nActuals'][k] + 0.5)
                    ws[f'L{i + 2}'] = int(dav_proj['As Sold Budget'][k] + 0.5)
                    ws[f'M{i + 2}'] = int(dav_proj['To Date\nActuals.1'][k] + 0.5)
            i += 1
            self.progress.emit(i)

        # Handle formatting
        thin = Side(border_style="thin", color="000000")
        not_center = [1, 2, 3, 11, 15, 18]
        for row in range(2, last_row + 1):
            ws[f'J{row}'] = f'=H{row}-I{row}'
            ws[f'N{row}'] = f'=L{row}-M{row}'
            ws[f'K{row}'] = f'=IFERROR(I{row}/H{row},"")'
            ws[f'O{row}'] = f'=IFERROR(M{row}/L{row},"")'
            ws[f'P{row}'] = f'=SUM(S{row}:AF{row})'
            ws[f'Q{row}'] = f'=SUM(S{row}:U{row})'
            ws[f'R{row}'] = f'=IFERROR(Q{row}/(I{row}+M{row}),"")'
            ws[f'K{row}'].number_format = '0%'
            ws[f'O{row}'].number_format = '0%'
            ws[f'R{row}'].number_format = '0.0%'
            if ws[f'A{row}'].value == 'DAV':
                ws[f'A{row}'].font = Font(color="ff0000")
            for col in range(1, 33):  # A-AF
                ws.cell(row=row, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col not in not_center:
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

        try:
            wb.save(proj_dash)
        except Exception:
            wb.close()
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {proj_dash} to allow editing", "Info", 0)
            return
        wb.close()
        self.progress.emit(0)
        self.finished.emit()
        ctypes.windll.user32.MessageBoxW(0, f"{proj_dash} update complete (Eng Hours)", "Info", 0)

    def UpdateEco(self):
        eco_folder_cov = Path("C:/RAJ_Vault/ECO/")
        eco_folder_dav = Path("C:/PT_Vault/ECO/")
        paco_folder_cov = Path("C:/RAJ_Vault/Planned Assy Change Orders/")
        paco_folder_dav = Path("C:/PT_Vault/Planned Assy Change Orders/")
        proj_dash = "C:/Users/Austin.Kidwell/Desktop/Project Dashboard R2.xlsx"
        #proj_dash = Path("//CPROME/Eng_Share/System Engineering/Visual Management/Test Folder/Project Dashboard R2.xlsx")

        # Setup macro for vault state
        cov_vault, dav_vault = "RAJ_Vault", "PT_Vault"
        pythoncom.CoInitialize()  # magic code to allow excel_app to work with threads
        excel_app = xw.App(visible=False)
        vb_wb = xw.Book("GetStatus.xlsm")
        macro1 = vb_wb.macro("Module1.GetSearch")

        dated_files_cov, dated_files_dav = [], []
        target_files_cov, target_files_dav = [], []

        start_search_date = datetime.datetime(2020, 1, 1).date()
        for file in os.listdir(eco_folder_cov):                         # get list of eco files
            if file.endswith("ECO.xlsm"):
                t = os.path.getmtime(eco_folder_cov / file)
                mod_date = datetime.datetime.fromtimestamp(t, tz=datetime.timezone.utc).date()
                if mod_date > start_search_date:                # get all files in folder after Jan 1, 2020
                    ftarget = os.path.join(eco_folder_cov, file)
                    dated_files_cov.append(ftarget)

        for root, dirs, files in os.walk(eco_folder_dav):
            for file in files:
                if file.endswith("ECO.xlsm"):
                    t = os.path.getmtime(root + "\\" + file)
                    mod_date = datetime.datetime.fromtimestamp(t, tz=datetime.timezone.utc).date()
                    if mod_date > start_search_date:                # get all files in folder after Jan 1, 2020
                        ftarget = os.path.join(root + "\\", file)
                        dated_files_dav.append(ftarget)

        for file in os.listdir(paco_folder_cov):                        # get list of paco files
            if file.startswith("PACO"):
                t = os.path.getmtime(paco_folder_cov / file)
                mod_date = datetime.datetime.fromtimestamp(t, tz=datetime.timezone.utc).date()
                if mod_date > start_search_date:                # get all files in folder after Jan 1, 2020
                    ftarget = os.path.join(paco_folder_cov, file)
                    dated_files_cov.append(ftarget)

        for file in os.listdir(paco_folder_dav):
            if file.startswith("PACO"):
                t = os.path.getmtime(paco_folder_dav / file)
                mod_date = datetime.datetime.fromtimestamp(t, tz=datetime.timezone.utc).date()
                if mod_date > start_search_date:                # get all files in folder after Jan 1, 2020
                    ftarget = os.path.join(paco_folder_dav, file)
                    dated_files_dav.append(ftarget)
        #print(target_files_cov)
        #print(len(dated_files_cov))
        #print(len(dated_files_dav))

        count = 0
        self.maximum.emit(len(dated_files_cov) + len(dated_files_dav))

        target_files_cov = macro1(cov_vault)   # only keep file with "ECO Complete" state
        target_files_cov = list(target_files_cov)
        # combine list by shared element (target files are sublists of dated files)
        target_files_cov = [x for x in dated_files_cov if any(b in x for b in target_files_cov)]
        count = len(dated_files_cov) - len(target_files_cov)
        self.progress.emit(count)

        target_files_dav = macro1(dav_vault)
        target_files_dav = list(target_files_dav)
        # combine list by shared element (target files are sublists of dated files)
        target_files_dav = [x for x in dated_files_dav if any(b in x for b in target_files_dav)]
        count += (len(dated_files_dav) - len(target_files_dav))
        self.progress.emit(count)

        vb_wb.app.quit()
        excel_app.kill()
        #print(len(target_files_cov))
        #print(len(target_files_dav))

        try:
            proj_df = pd.read_excel(proj_dash, sheet_name="Projects", header=[0])
        except Exception:
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {proj_dash}", "Info", 0)
            return

        mask = proj_df['Site'].str.contains('COV', na=False, regex=True)
        proj_df_cov = proj_df[mask]
        mask = proj_df['Site'].str.contains('DAV', na=False, regex=True)
        proj_df_dav = proj_df[mask]
        #print(proj_df_cov)
        #print(proj_df_dav)

        def get_eco(x):      # get absence type based on row num
            return {
                "DESIGN CHECKING ERROR (JDC)": 1,
                "DESIGN CONCEPT ERROR (JDC)": 2,
                "SAFETY (JDC)": 3,
                "PROJECT SCOPE CHANGE (JDC)": 4,
                "NON-ENGINEERING REQUEST (JDC)": 5,
                "LEGACY ERROR (JDC)": 6,
                "OBSOLESCENCE (JDC)": 7,
                "OBSOLESCENCE (SDC)": 8,
                "COST REDUCTION (SDC)": 9,
                "ALTERNATE CONSTRUCTION (SDC)": 10,
                "FEATURE ADDITION (SDC)": 11,
                "INCREASE FUTURE MARKETABILITY (SDC)": 11,
                "SAFETY (SDC)": 12,
                "LEGACY ERROR (SDC)": 13
            }.get(x, 14)

        proj_list_cov = list(proj_df_cov["Project Number"])
        proj_list_dav = list(proj_df_dav["Project Number"])

        broken = Path("C:/RAJ_Vault/ECO/0VPK01670 REV 2 ECO.xlsm")
        remove_last = Path("C:/RAJ_Vault/ECO/28LMC730117534 REV 2 ECO.xlsm")
        find_eco_cov = []
        find_eco_dav = []
        for file in target_files_cov:   # search COV eco vault to add eco of active projects
            #print(file)
            if os.path.samefile(file, broken) or 'test' in file.lower():
                catg, proj = "", ""
            else:
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                try:
                    ws = wb['ECO Form Sheet 1']
                    catg = ws["B13"].value
                    if ws["K13"].value:
                        proj = str(ws["K13"].value).strip()
                    else:
                        proj = ''
                    if os.path.samefile(file, remove_last):
                        proj = proj[:10]
                except:
                    ws = wb['PACO Form Sheet 1']
                    catg = "paco"
                    if ws["A8"].value:
                        proj = str(ws["A8"].value).strip()
                    else:
                        proj = str(ws["C8"].value).strip()
            #print(count, f':{proj}:{catg}:{len(proj)}')
            if len(proj) == 5:
                if proj > '17000':
                    proj = 'C009.' + proj
                else:
                    proj = 'R009.' + proj
            if len(proj) == 10:
                proj = proj[0].upper() + proj[1:10]
            if proj in proj_list_cov:
                find_eco_cov.append((proj, catg))
            elif ',' or ' ' or '\n' or '/' in proj:
                proj = proj.replace('/', '-')
                proj = proj.replace('"', '')
                multi_proj = re.split(',|, | |\n|;|\*|&|\. |\t', proj)
                for item in multi_proj:
                    if '-' in item:
                        start, end = item.split('-')
                        try:
                            start = start.split('.')[1]
                            if len(start) > len(end):
                                longer = len(start) - len(end)
                                end = start[0:longer] + end
                            start, end = int(start), int(end)
                            while start != end:
                                temp = '00000' + str(start)
                                multi_proj.append(temp[-5:])
                                start += 1
                        except:
                            pass
                        finally:
                            item = '00000' + str(end)
                            item = item[-5:]
                    if len(item) > 5:
                        item = item[-5:]
                    if len(item) == 5:
                        if item > '17000':
                            item = 'C009.' + item
                        else:
                            item = 'R009.' + item
                    if item in proj_list_cov:
                        find_eco_cov.append((item, catg))
            count += 1
            self.progress.emit(count)

        for file in target_files_dav:   # search DAV eco vault to add eco of active projects
            #print(file)
            if 'test' in file.lower():
                catg, proj = "", ""
            else:
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                try:
                    ws = wb['ECO Form Sheet 1']
                    catg = ws["B13"].value
                    if ws["K13"].value:
                        proj = str(ws["K13"].value).strip()
                    else:
                        proj = ''
                except:
                    ws = wb['PACO Form Sheet 1']
                    catg = "paco"
                    if ws["A8"].value:
                        proj = str(ws["A8"].value).strip()
                    else:
                        proj = str(ws["C8"].value).strip()
            #print(count, f':{proj}:{catg}:{len(proj)}')
            if len(proj) == 5:
                if proj > '17000':
                    proj = 'C009.' + proj
                else:
                    proj = 'R009.' + proj
            if len(proj) == 10:
                proj = proj[0].upper() + proj[1:10]
            if proj in proj_list_dav:
                find_eco_dav.append((proj, catg))
            elif ',' or ' ' or '\n' or '/'  in proj:
                proj = proj.replace('/', '-')
                multi_proj = re.split(',|, | |\n|;|\*|&|\. ', proj)
                for item in multi_proj:
                    if '-' in item:
                        start, end = item.split('-')
                        try:
                            start = start.split('.')[1]
                            if len(start) > len(end):
                                longer = len(start) - len(end)
                                end = start[0:longer] + end
                            start, end = int(start), int(end)
                            while start != end:
                                temp = '00000'+str(start)
                                multi_proj.append(temp[-5:])
                                start += 1
                        except:
                            pass
                        finally:
                            item = '00000'+str(end)
                            item = item[-5:]
                    if len(item) > 5:
                        item = item[-5:]
                    if len(item) == 5:
                        if item > '17000':
                            item = 'C009.' + item
                        else:
                            item = 'R009.' + item
                    if item in proj_list_dav:
                        find_eco_dav.append((item, catg))
            count += 1
            self.progress.emit(count)

        wb.close()
        count_eco_cov = [[0 for x in range(15)] for y in range(len(proj_list_cov))] # blank matrix to store eco counts
        proj_loc_cov = {}
        for i in range(len(proj_list_cov)):     # get dictionary to hold row index of COV projects
            count_eco_cov[i][0] = proj_list_cov[i]
            proj_loc_cov[proj_list_cov[i]] = i

        for i in range(len(find_eco_cov)):  # count COV eco reports
            row = proj_loc_cov[find_eco_cov[i][0]]
            col = get_eco(find_eco_cov[i][1])
            count_eco_cov[row][col] += 1

        count_eco_dav = [[0 for x in range(15)] for y in range(len(proj_list_dav))] # blank matrix to store eco counts
        proj_loc_dav = {}
        for i in range(len(proj_list_dav)):     # get dictionary to hold row index of DAV projects
            count_eco_dav[i][0] = proj_list_dav[i]
            proj_loc_dav[proj_list_dav[i]] = i

        for i in range(len(find_eco_dav)):  # count DAV eco reports
            row = proj_loc_dav[find_eco_dav[i][0]]
            col = get_eco(find_eco_dav[i][1])
            count_eco_dav[row][col] += 1

        dav_start = len(count_eco_cov) + 2
        wb = openpyxl.load_workbook(proj_dash)
        ws = wb['Projects']
        for i in range(len(count_eco_cov)):     # populate COV eco numbers
            ws[f'S{i + 2}'], ws[f'T{i + 2}'] = count_eco_cov[i][1], count_eco_cov[i][2]
            ws[f'U{i + 2}'], ws[f'V{i + 2}'] = count_eco_cov[i][3], count_eco_cov[i][4]
            ws[f'W{i + 2}'], ws[f'X{i + 2}'] = count_eco_cov[i][5], count_eco_cov[i][6]
            ws[f'Y{i + 2}'], ws[f'Z{i + 2}'] = count_eco_cov[i][7], count_eco_cov[i][8]
            ws[f'AA{i + 2}'], ws[f'AB{i + 2}'] = count_eco_cov[i][9], count_eco_cov[i][10]
            ws[f'AC{i + 2}'], ws[f'AD{i + 2}'] = count_eco_cov[i][11], count_eco_cov[i][12]
            ws[f'AE{i + 2}'], ws[f'AF{i + 2}'] = count_eco_cov[i][13], count_eco_cov[i][14]

        for i in range(len(count_eco_dav)):     # populate DAV eco numbers
            ws[f'S{i + dav_start}'], ws[f'T{i + dav_start}'] = count_eco_dav[i][1], count_eco_dav[i][2]
            ws[f'U{i + dav_start}'], ws[f'V{i + dav_start}'] = count_eco_dav[i][3], count_eco_dav[i][4]
            ws[f'W{i + dav_start}'], ws[f'X{i + dav_start}'] = count_eco_dav[i][5], count_eco_dav[i][6]
            ws[f'Y{i + dav_start}'], ws[f'Z{i + dav_start}'] = count_eco_dav[i][7], count_eco_dav[i][8]
            ws[f'AA{i + dav_start}'], ws[f'AB{i + dav_start}'] = count_eco_dav[i][9], count_eco_dav[i][10]
            ws[f'AC{i + dav_start}'], ws[f'AD{i + dav_start}'] = count_eco_dav[i][11], count_eco_dav[i][12]
            ws[f'AE{i + dav_start}'], ws[f'AF{i + dav_start}'] = count_eco_dav[i][13], count_eco_dav[i][14]

        try:
            wb.save(proj_dash)
        except Exception:
            wb.close()
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {proj_dash} to allow editing", "Info", 0)
            return
        wb.close()
        self.progress.emit(0)
        self.finished.emit()
        ctypes.windll.user32.MessageBoxW(0, f"{proj_dash} update complete (ECO)", "Info", 0)

    def ReleaseData(self):
        today = datetime.date.today()                   # Initialize current date
        this_year = today.year
        this_month = today.month

        steering_boards_cov = Path("//uscoit-sv006/Public/Steering Boards/")
        steering_boards_dav = Path("C:/PT_Vault/Steering Boards/")

        proj_dash = "C:/Users/Austin.Kidwell/Desktop/Project Dashboard R2.xlsx"
        #proj_dash = Path("//CPROME/Eng_Share/System Engineering/Visual Management/Test Folder/Project Dashboard R2.xlsx")

        cov_files = []
        dav_files = []
        for file in os.listdir(steering_boards_cov):
            if file.endswith(".xlsm") and "~$" not in file:
                ftarget = os.path.join(steering_boards_cov, file)
                cov_files.append(ftarget)

        for file in os.listdir(steering_boards_dav):
            if file.endswith(".xlsm") and "~$" not in file:
                ftarget = os.path.join(steering_boards_dav, file)
                dav_files.append(ftarget)

        #print(len(cov_files))
        #print(len(dav_files))
        temp, temp2 = [], []
        for file in cov_files:       # Double-check only C009 and R009 projects
            if 'C009' in file or 'R009' in file:
                temp.append(file)
        cov_files = temp
        for file in dav_files:       # Double-check only C009 and R009 projects
            if 'C009' in file or 'R009' in file:
                temp2.append(file)
        dav_files = temp2
        # Mimic excel default cond_format: breakpoints=[0.85, 0.89, 0.92, 0.95], color=[(248, 105, 107),
                                    # (251, 170, 119), (255, 235, 132), (177, 213, 128), (99, 190, 123)]):
        # get conditional format for range
        def cond_format(x, breakpoints=[0.90, 0.95], color=[(248, 105, 107), (255, 235, 132), (99, 190, 123)]):
            if x is None:
                return None
            i = bisect(breakpoints, x)
            return color[i]
        # Only read C009 and R009 files
        covReleaseData, davReleaseData, covDueIn2Weeks, davDueIn2Weeks, covEngDash, davEngDash = [], [], [], [], [], []
        count = 0
        self.maximum.emit(len(cov_files) + len(dav_files) + 3)
        for file in cov_files:              # read COV steering boards data
            root, sn = file.rsplit('\\', 1)
            strSN = sn[0:10]
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb['Steering Board']
            strCustomer = ws["B2"].value
            strMachType = ws["B3"].value
            ME_FinalRelDate, EE_FinalRelDate = "", ""
            ME_TotAssys, EE_TotAssys, ME_RelOnTime, EE_RelOnTime, ME_RelLate, EE_RelLate = 0, 0, 0, 0, 0, 0
            ME_OpenAssy, EE_OpenAssy, ME_DueNextWeek, EE_DueNextWeek, ME_DueTwoWeeks, EE_DueTwoWeeks = 0, 0, 0, 0, 0, 0
            last_row = 0
            for row in ws.iter_rows(max_row=400):
                if not all([cell.value == None for cell in row]):
                    last_row += 1
            #print(last_row)
            for i in range(7, last_row + 1):
                strDesc = ws[f"C{i}"].value
                strCat = ws[f"F{i}"].value
                strEng = ws[f"G{i}"].value
                dateDue = ws[f"K{i}"].value
                dateRelease = ws[f"L{i}"].value
                expDateRelease = ws[f"M{i}"].value
                if dateRelease is None:                         # stops crash when dateRelease doesn't exist
                    pass
                elif this_year == dateRelease.year and this_month == dateRelease.month: # get release data
                    if expDateRelease is None:
                        onTime = "TRUE"
                        daysLate = None
                    elif expDateRelease >= dateRelease:
                        onTime = "TRUE"
                        daysLate = None
                    else:
                        onTime = "FALSE"
                        daysLate = np.busday_count(expDateRelease.date(), dateRelease.date())
                    if dateDue is not None:
                        dateDue = datetime.datetime.strptime(str(dateDue.date()), '%Y-%m-%d').strftime('%m/%d/%Y')
                    dateRelease = datetime.datetime.strptime(str(dateRelease.date()), '%Y-%m-%d').strftime('%m/%d/%Y')
                    if expDateRelease is not None:
                        expDateRelease = datetime.datetime.strptime(str(expDateRelease.date()), '%Y-%m-%d').strftime(
                            '%m/%d/%Y')
                    covReleaseData.append(("COV", strSN, strCustomer, strMachType, strDesc, strCat, strEng, dateDue,
                                            dateRelease, expDateRelease, onTime, daysLate))
                if expDateRelease == 'TBD':
                    expDateRelease = None
                if dateRelease is None and ws[f"W{i}"].value is not None:               # get due in 2 weeks data
                    if expDateRelease is None:
                        daysLeft = None
                        expDateRelease = "Missing"
                        covDueIn2Weeks.append(("COV", strSN, strCustomer, strMachType, strDesc, strCat, strEng,
                                               expDateRelease, daysLeft))
                        expDateRelease = None
                    elif (expDateRelease.date() - today).days < 14:
                        daysLeft = (expDateRelease.date() - today).days
                        expDateRelease = datetime.datetime.strptime(str(expDateRelease.date()), '%Y-%m-%d').strftime(
                            '%m/%d/%Y')
                        covDueIn2Weeks.append(("COV", strSN, strCustomer, strMachType, strDesc, strCat, strEng,
                                            expDateRelease, daysLeft))
                if strCat == "Mechanical" and (dateRelease is not None or expDateRelease is not None): # get eng dash
                    if dateRelease is not None:
                        if type(dateRelease) is str:
                            dateRelease = datetime.datetime.strptime(dateRelease, '%m/%d/%Y')
                    ME_TotAssys += 1
                    if type(expDateRelease) is str:
                        expDateRelease = datetime.datetime.strptime(expDateRelease, '%m/%d/%Y')
                    if expDateRelease is not None and (ME_FinalRelDate == "" or expDateRelease > ME_FinalRelDate):
                        ME_FinalRelDate = expDateRelease
                    if dateRelease is None:
                        ME_OpenAssy += 1
                        if (expDateRelease.date() - today).days < 7:
                            ME_DueNextWeek += 1
                        if (expDateRelease.date() - today).days < 14:
                            ME_DueTwoWeeks += 1
                    elif expDateRelease is not None and dateRelease > expDateRelease:
                        ME_RelLate += 1
                    else:
                        ME_RelOnTime += 1
                elif strCat == "Electrical" and (dateRelease is not None or expDateRelease is not None):
                    if dateRelease is not None:
                        if type(dateRelease) is str:
                            dateRelease = datetime.datetime.strptime(dateRelease, '%m/%d/%Y')
                    EE_TotAssys += 1
                    if type(expDateRelease) is str:
                        expDateRelease = datetime.datetime.strptime(expDateRelease, '%m/%d/%Y')
                    if expDateRelease is not None and (EE_FinalRelDate == "" or expDateRelease > EE_FinalRelDate):
                        EE_FinalRelDate = expDateRelease
                    if dateRelease is None:
                        EE_OpenAssy += 1
                        if (expDateRelease.date() - today).days < 7:
                            EE_DueNextWeek += 1
                        if (expDateRelease.date() - today).days < 14:
                            EE_DueTwoWeeks += 1
                    elif expDateRelease is not None and dateRelease > expDateRelease:
                        EE_RelLate += 1
                    else:
                        EE_RelOnTime += 1
            covEngDash.append(("COV", strSN, f"{strCustomer} ({strMachType})", "ME", ME_FinalRelDate, ME_TotAssys,
                               ME_RelOnTime, ME_RelLate, ME_OpenAssy, ME_DueNextWeek, ME_DueTwoWeeks,
                               f'=IFERROR({ME_RelOnTime}/({ME_RelOnTime}+{ME_RelLate}),"")', "EE", EE_FinalRelDate,
                               EE_TotAssys, EE_RelOnTime, EE_RelLate, EE_OpenAssy, EE_DueNextWeek, EE_DueTwoWeeks,
                               f'=IFERROR({EE_RelOnTime}/({EE_RelOnTime}+{EE_RelLate}),"")'))
            count += 1
            self.progress.emit(count)
        wb.close()
        for file in dav_files:              # get DAV steering board data
            root, sn = file.rsplit('\\', 1)
            strSN = sn[0:10]
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb['Steering Board']
            strCustomer = ws["B2"].value
            strMachType = ws["B3"].value
            ME_FinalRelDate, EE_FinalRelDate = "", ""
            ME_TotAssys, EE_TotAssys, ME_RelOnTime, EE_RelOnTime, ME_RelLate, EE_RelLate = 0, 0, 0, 0, 0, 0
            ME_OpenAssy, EE_OpenAssy, ME_DueNextWeek, EE_DueNextWeek, ME_DueTwoWeeks, EE_DueTwoWeeks = 0, 0, 0, 0, 0, 0
            last_row = 0
            for row in ws.iter_rows(max_row=400):
                if not all([cell.value == None for cell in row]):
                    last_row += 1
            for i in range(7, last_row + 1):
                strDesc = ws[f"C{i}"].value
                strCat = ws[f"F{i}"].value
                strEng = ws[f"G{i}"].value
                dateDue = ws[f"K{i}"].value
                dateRelease = ws[f"L{i}"].value
                expDateRelease = ws[f"K{i}"].value
                if dateRelease is None:                         # stops crash when dateRelease doesn't exist
                    pass
                elif this_year == dateRelease.year and this_month == dateRelease.month:     # get release data
                    if expDateRelease is None:
                        onTime = "TRUE"
                        daysLate = None
                    elif expDateRelease >= dateRelease:
                        onTime = "TRUE"
                        daysLate = None
                    else:
                        onTime = "FALSE"
                        daysLate = np.busday_count(expDateRelease.date(), dateRelease.date())
                    if dateDue is not None:
                        dateDue = datetime.datetime.strptime(str(dateDue.date()), '%Y-%m-%d').strftime('%m/%d/%Y')
                    dateRelease = datetime.datetime.strptime(str(dateRelease.date()), '%Y-%m-%d').strftime('%m/%d/%Y')
                    if expDateRelease is not None:
                        expDateRelease = datetime.datetime.strptime(str(expDateRelease.date()), '%Y-%m-%d').strftime(
                            '%m/%d/%Y')
                    davReleaseData.append(("DAV", strSN, strCustomer, strMachType, strDesc, strCat, strEng, dateDue,
                                            dateRelease, expDateRelease, onTime, daysLate))
                if expDateRelease == 'TBD':
                    expDateRelease = None
                if dateRelease is None and ws[f"W{i}"].value is not None:      # get due in 2 weeks data
                    if expDateRelease is None:
                        daysLeft = None
                        expDateRelease = "Missing"
                        davDueIn2Weeks.append(("DAV", strSN, strCustomer, strMachType, strDesc, strCat, strEng,
                                               expDateRelease, daysLeft))
                        expDateRelease = None
                    elif (expDateRelease.date() - today).days < 14:
                        daysLeft = (expDateRelease.date() - today).days
                        expDateRelease = datetime.datetime.strptime(str(expDateRelease.date()), '%Y-%m-%d').strftime(
                            '%m/%d/%Y')
                        davDueIn2Weeks.append(("DAV", strSN, strCustomer, strMachType, strDesc, strCat, strEng,
                                            expDateRelease, daysLeft))
                if strCat == "Mechanical" and (dateRelease is not None or expDateRelease is not None): # get eng dash
                    if dateRelease is not None:
                        if type(dateRelease) is str:
                            dateRelease = datetime.datetime.strptime(dateRelease, '%m/%d/%Y')
                    ME_TotAssys += 1
                    if type(expDateRelease) is str:
                        expDateRelease = datetime.datetime.strptime(expDateRelease, '%m/%d/%Y')
                    if expDateRelease is not None and (ME_FinalRelDate == "" or expDateRelease > ME_FinalRelDate):
                        ME_FinalRelDate = expDateRelease
                    if dateRelease is None:
                        ME_OpenAssy += 1
                        if (expDateRelease.date() - today).days < 7:
                            ME_DueNextWeek += 1
                        if (expDateRelease.date() - today).days < 14:
                            ME_DueTwoWeeks += 1
                    elif expDateRelease is not None and dateRelease > expDateRelease:
                        ME_RelLate += 1
                    else:
                        ME_RelOnTime += 1
                elif strCat == "Electrical" and (dateRelease is not None or expDateRelease is not None):
                    if dateRelease is not None:
                        if type(dateRelease) is str:
                            dateRelease = datetime.datetime.strptime(dateRelease, '%m/%d/%Y')
                    EE_TotAssys += 1
                    if type(expDateRelease) is str:
                        expDateRelease = datetime.datetime.strptime(expDateRelease, '%m/%d/%Y')
                    if expDateRelease is not None and (EE_FinalRelDate == "" or expDateRelease > EE_FinalRelDate):
                        EE_FinalRelDate = expDateRelease
                    if dateRelease is None:
                        EE_OpenAssy += 1
                        if (expDateRelease.date() - today).days < 7:
                            EE_DueNextWeek += 1
                        if (expDateRelease.date() - today).days < 14:
                            EE_DueTwoWeeks += 1
                    elif expDateRelease is not None and dateRelease > expDateRelease:
                        EE_RelLate += 1
                    else:
                        EE_RelOnTime += 1
            davEngDash.append(("DAV", strSN, f"{strCustomer} ({strMachType})", "ME", ME_FinalRelDate, ME_TotAssys,
                               ME_RelOnTime, ME_RelLate, ME_OpenAssy, ME_DueNextWeek, ME_DueTwoWeeks,
                               f'=IFERROR({ME_RelOnTime}/({ME_RelOnTime}+{ME_RelLate}),"")', "EE", EE_FinalRelDate,
                               EE_TotAssys, EE_RelOnTime, EE_RelLate, EE_OpenAssy, EE_DueNextWeek, EE_DueTwoWeeks,
                               f'=IFERROR({EE_RelOnTime}/({EE_RelOnTime}+{EE_RelLate}),"")'))
            count += 1
            self.progress.emit(count)
        wb.close()

        releaseData = covReleaseData + davReleaseData
        dueIn2Weeks = covDueIn2Weeks + davDueIn2Weeks
        EngDash = covEngDash + davEngDash

        pythoncom.CoInitialize()     # magic code to allow excel_app to work with threads
        excel_app = xw.App(visible=False)
        self.progress.emit(count + 1)
        try:
            wb = xw.Book(proj_dash)
        except Exception:
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Can't locate file {proj_dash}", "Info", 0)
            return
        ws = wb.sheets['Release Data']
        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        ws.range(f"A2:L{max_row}").clear_contents()
        for i in range(len(releaseData)):           # populate release data tab
            ws.range(f"A{i + 2}").value = releaseData[i][0]
            ws.range(f"B{i + 2}").value = releaseData[i][1]
            ws.range(f"C{i + 2}").value = releaseData[i][2]
            ws.range(f"D{i + 2}").value = releaseData[i][3]
            ws.range(f"E{i + 2}").value = releaseData[i][4]
            ws.range(f"F{i + 2}").value = releaseData[i][5]
            ws.range(f"G{i + 2}").value = releaseData[i][6]
            ws.range(f"H{i + 2}").value = releaseData[i][7]
            ws.range(f"I{i + 2}").value = releaseData[i][8]
            ws.range(f"J{i + 2}").value = releaseData[i][9]
            ws.range(f"K{i + 2}").value = releaseData[i][10]
            ws.range(f"L{i + 2}").value = releaseData[i][11]
            if ws.range(f'A{i + 2}').value == 'DAV':
                ws.range(f'A{i + 2}').api.Font.ColorIndex = 3 #'ff0000'
            else:
                ws.range(f'A{i + 2}').api.Font.ColorIndex = 0 #'000000'

        self.progress.emit(count + 2)
        ws = wb.sheets['Due in 2 Weeks']
        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        ws[f"A2:H{max_row}"].clear_contents()
        ws[f"I2:I{max_row}"].clear()
        for j in range(len(dueIn2Weeks)):       # populate due in 2 weeks tab
            ws.range(f"A{j + 2}").value = dueIn2Weeks[j][0]
            ws.range(f"B{j + 2}").value = dueIn2Weeks[j][1]
            ws.range(f"C{j + 2}").value = dueIn2Weeks[j][2]
            ws.range(f"D{j + 2}").value = dueIn2Weeks[j][3]
            ws.range(f"E{j + 2}").value = dueIn2Weeks[j][4]
            ws.range(f"F{j + 2}").value = dueIn2Weeks[j][5]
            ws.range(f"G{j + 2}").value = dueIn2Weeks[j][6]
            ws.range(f"H{j + 2}").value = dueIn2Weeks[j][7]
            ws.range(f"I{j + 2}").value = dueIn2Weeks[j][8]
            if ws.range(f'A{j + 2}').value == 'DAV':
                ws.range(f'A{j + 2}').api.Font.ColorIndex = 3 #'ff0000'
            else:
                ws.range(f'A{j + 2}').api.Font.ColorIndex = 0 #'000000'
            if ws.range(f"I{j + 2}").value is None or ws.range(f"I{j + 2}").value >= 0:
                ws.range(f"I{j + 2}").color = (255, 255, 0)
            else:
                ws.range(f"I{j + 2}").color = (255, 0, 0)

        self.progress.emit(count + 3)
        shrink = False
        ws = wb.sheets['Engineering Dashboard']
        max_row = ws.range('A' + str(ws.cells.last_cell.row)).end('up').row
        ws[f"A4:U{max_row}"].clear_contents()
        for k in range(len(EngDash)):       # populate eng dashboard tab
            ws.range(f"A{k + 4}").value = EngDash[k][0]
            ws.range(f"B{k + 4}").value = EngDash[k][1]
            ws.range(f"C{k + 4}").value = EngDash[k][2]
            ws.range(f"D{k + 4}").value = EngDash[k][3]
            ws.range(f"E{k + 4}").value = EngDash[k][4]
            ws.range(f"F{k + 4}").value = EngDash[k][5]
            ws.range(f"G{k + 4}").value = EngDash[k][6]
            ws.range(f"H{k + 4}").value = EngDash[k][7]
            ws.range(f"I{k + 4}").value = EngDash[k][8]
            ws.range(f"J{k + 4}").value = EngDash[k][9]
            ws.range(f"K{k + 4}").value = EngDash[k][10]
            ws.range(f"L{k + 4}").value = EngDash[k][11]
            ws.range(f"L{k + 4}").value = ws.range(f"L{k + 4}").value
            ws.range(f"M{k + 4}").value = EngDash[k][12]
            ws.range(f"N{k + 4}").value = EngDash[k][13]
            ws.range(f"O{k + 4}").value = EngDash[k][14]
            ws.range(f"P{k + 4}").value = EngDash[k][15]
            ws.range(f"Q{k + 4}").value = EngDash[k][16]
            ws.range(f"R{k + 4}").value = EngDash[k][17]
            ws.range(f"S{k + 4}").value = EngDash[k][18]
            ws.range(f"T{k + 4}").value = EngDash[k][19]
            ws.range(f"U{k + 4}").value = EngDash[k][20]
            ws.range(f"U{k + 4}").value = ws.range(f"U{k + 4}").value
            if ws.range(f'A{k + 4}').value == 'DAV':
                ws.range(f'A{k + 4}').api.Font.ColorIndex = 3
            else:
                ws.range(f'A{k + 4}').api.Font.ColorIndex = 0
            ws.range(f"L{k + 4}").color = cond_format(ws.range(f"L{k + 4}").value)
            ws.range(f"U{k + 4}").color = cond_format(ws.range(f"U{k + 4}").value)
        new_max = len(EngDash) + 3
        if new_max < max_row:
            ws.range(f"A{new_max + 1}:U{max_row}").clear()
            ws.activate()
            #tab = xw.Range('Table2[#All]').resize(row_size=len(EngDash) + 1)
            #ws.tables["Table2"].resize(tab)
            shrink = True
        elif new_max > max_row:
            light_blue = (217, 225, 242)
            blue = (180, 198, 231)
            for row in range(max_row + 1, new_max + 1):
                if row % 2 == 1:
                    ws.range(f"M{row}:T{row}").color = light_blue
                    if ws.range(f"U{row}").value is None:
                        ws.range(f"U{row}").color = light_blue
                else:
                    ws.range(f"M{row}:T{row}").color = blue
                    if ws.range(f"U{row}").value is None:
                        ws.range(f"U{row}").color = blue
            ws.range(f"D{max_row}:D{new_max}").api.Borders(12).LineStyle = 1
            ws.range(f"D{max_row}:D{new_max}").api.Borders(12).Weight = 3
            ws.range(f"M{max_row}:M{new_max}").api.Borders(12).LineStyle = 1
            ws.range(f"M{max_row}:M{new_max}").api.Borders(12).Weight = 3
            ws.range(f"D{max_row}:D{new_max}").api.HorizontalAlignment = xw.constants.HAlign.xlHAlignCenter

        try:
            wb.save(proj_dash)
        except Exception:
            excel_app.kill()
            self.progress.emit(0)
            self.finished.emit()
            pythoncom.CoUninitialize()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {proj_dash} to allow editing", "Info", 0)
            return
        excel_app.kill()
        pythoncom.CoUninitialize()
        try:
            if shrink == True:
                wb1 = openpyxl.load_workbook(proj_dash)  # Make excel sheet ready to edit
                ws = wb1['Engineering Dashboard']
                tab = ws.tables["Table2"]
                tab.ref = f"A3:U{len(EngDash) + 3}"
                wb1.save(proj_dash)
        except Exception:
            self.progress.emit(0)
            self.finished.emit()
            ctypes.windll.user32.MessageBoxW(0, f"Close excel file {proj_dash} to allow editing", "Info", 0)
            return
        self.progress.emit(0)
        self.finished.emit()
        ctypes.windll.user32.MessageBoxW(0, f"{proj_dash} update complete (Release Data)", "Info", 0)
